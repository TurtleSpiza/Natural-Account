#!/usr/bin/env python3
"""Extend 00_Running_Transaction_Listing.xlsx for the five 11-Jun handover
workstreams (73211, 73541, 7B532, 7B214, 62121/62125). One-shot session script:
adds five line tabs in the build_listing.py rendering convention (22 columns
including Journal) and rewrites the Summary tab rows and TOTAL (live SUM).
Verdicts come from the verification records and NA73211_Evidence_Manifest.csv
only; no recode dollar is populated (candidates held un-instructed)."""
import csv, json, re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY = "1F4E79"
HDR = Font(name="Segoe UI", bold=True, color="FFFFFF", size=10)
B = Font(name="Segoe UI", size=10)
band = PatternFill("solid", fgColor="F2F2F2")
RAG = {'G': ("C6EFCE", "006100"), 'A': ("FFEB9C", "9C6500"),
       'R': ("FFC7CE", "9C0006"), 'P': ("D9D9D9", "595959"),
       'J': ("A9D08E", "375623")}
thin = Side(style="thin", color="D9D9D9")
bd = Border(thin, thin, thin, thin)
ctr = Alignment(horizontal="center", vertical="center")
lft = Alignment(horizontal="left", vertical="center")
COLS = [("Account", 8), ("Svc", 6), ("PK", 10), ("Section", 22), ("Period", 6),
        ("Date", 11), ("Reference", 13), ("Vendor", 24), ("Cardholder", 15),
        ("Details", 44), ("ExGST", 10), ("InclX11", 10), ("Attach", 6),
        ("L1 Svc/PK", 9), ("L1 Emp", 8), ("L2 NA", 7), ("L3 Evid", 8),
        ("L4 Tax", 7), ("Overall", 8), ("Finding / note", 66), ("Recode To", 14),
        ("Journal", 30)]
RAGCOLS = {"L1 Svc/PK", "L1 Emp", "L2 NA", "L3 Evid", "L4 Tax", "Overall"}

svc_map = json.load(open('.claude/skills/lcc-coding-review/svc_map.json'))['service_codes']
def svc_info(s):
    r = svc_map.get(s)
    return (r.get('pk') or '', r.get('section') or '') if r else ('', '')

EXP = '_Sources_11-Jun-2026/Handover_Ledger_Exports_11-Jun-2026/Ledger Accounts Transactions Table - '

def load(path):
    raw = pd.read_excel(path, header=None, engine='calamine')
    hdr = next(i for i, r in raw.iterrows() if any('Short Description' in str(v) for v in r.values))
    df = raw.iloc[hdr + 1:].copy()
    df.columns = raw.iloc[hdr].tolist()
    df = df[df['Reference'].notna()].copy()
    df['amt'] = pd.to_numeric(df['Transaction Amount'], errors='coerce')
    return df

# ---------------- 73211 classification (from the Verification Record addendum
# of 11-Jun-2026 and the canonical evidence manifest) ----------------
man = list(csv.DictReader(open('02_NA73211_Minor_Contracts/NA73211_Evidence_Manifest.csv')))
SIGHTED_REFS = set()
for r in man:
    if r['file'] == 'GAP':
        continue
    lr = r['line_ref']
    m = re.match(r'GJ078960 \(EWN (\d+)\)', lr)
    SIGHTED_REFS.add(('GJ078960', m.group(1)) if m else lr)

VENDOR = {}
for ref in ['2907532', '2907801', '2908018', '2908241', '2908351', '2908470',
            '2908632', '2908653', '2908664', '2908883', '2906937', '2908705', '2907368', '2907673']:
    VENDOR[ref] = 'Lockwise Locksmiths' if ref not in ('2906937', '2908705', '2907368', '2907673') else ''
for ref in ['49359879350725', '49359879350825', '49359879350925', '49359879351025',
            '49359879351125', '49359879351225', '49359879350126', '49359879350226',
            '49359879350326', '49359879350426', '49359879350526']:
    VENDOR[ref] = 'Telstra'
VENDOR.update({
    'INV-0352': 'Destination Trails Pty Ltd', 'INV-0354': 'Destination Trails Pty Ltd',
    'INV-0355': 'Destination Trails Pty Ltd', 'GJ078945': 'Destination Trails Pty Ltd',
    'GJ079117': 'Destination Trails Pty Ltd',
    'INV-0785': 'Chronicle Rip', 'INV-0856': 'Chronicle Rip', 'INV-0868': 'Chronicle Rip',
    'INV-0873': 'Chronicle Rip',
    'LCC31326': 'INCON Solutions T/A Edenstone Masonry NQ',
    'INV-5747': 'Reactive Generators', 'INV-0087': 'Firesight', 'INV-5888': 'Play Force',
    '27168': 'Tennyson Group', 'INV-1917': 'CAASie', '18970': 'RST Systems',
    'IAS21464': 'Independent Arboricultural Services', '39657': 'IPWEA',
    '00001400': 'PSIA', 'RL195364': 'Empire Office Furniture',
    'DOC0079602': 'DDLS T/A Lumify', 'SC10918': 'Worssell',
    'TE004838': 'Sinch MessageMedia', 'INV-251010(2)': 'Organisational Risk Consulting',
    'INV-13157': 'Bendelta',
    '3223657': 'T2 Electrical', '3223717': 'T2 Electrical', '3222869': 'T2 Electrical',
    '3223786': 'T2 Electrical',
    'TBM 00097461': 'TBM', 'TBM 00097583': 'TBM', 'TBM 00097626': 'TBM', 'TBM 00097780': 'TBM',
})

L1_AMBER = {'GJ079190', 'GJ079210', 'INV-5888', 'SC10918', 'TE004838'}
# F2 / F3 / F4 / new-candidate lines: L2 AMBER, candidate destination in the note only
L2_AMBER_NOTE = {
    'LCC31326': 'F2 candidate 73211>73212 (above the $20,000 floor on its face); INCON ABR sighted 11-Jun (active, GST reg; NOTE-tier young-entity concerns). New reading: materials supply, not contract services - arguably a stores/materials question. Held un-instructed.',
    'INV-0352': 'F2 candidate 73211>73212 (Destination Trails series $43,114 fully evidenced); split pair GJ078945/GJ079117 recomputed, nets $0.00, direction correct. Held un-instructed.',
    'INV-0354': 'F2 candidate 73211>73212 (Destination Trails series). Held un-instructed.',
    'INV-0355': 'F2 candidate 73211>73212 (Destination Trails series). Held un-instructed.',
    'INV-0785': 'F2 candidate 73211>73212 (Chronicle series $29,042.45). Map edits, Beenleigh monumental s5. Held un-instructed.',
    'INV-0856': 'F2 candidate 73211>73212 (Chronicle series). Aerial imagery; arguably survey services. Held un-instructed.',
    'INV-0868': 'F2 candidate plus NEW: 12-month software licence renewal (7 cemeteries) - 73566 candidate mirroring the 73564 ruling. Held un-instructed.',
    'INV-0873': 'F2 candidate 73211>73212 (Chronicle series; arguably above-threshold alone). 360 mapping, four cemeteries; arguably survey services. Held un-instructed.',
    'INV-0087': 'NEW candidate 73211>73601: bushfire risk analysis with options assessment; four-element consultancy test reads as met. Held un-instructed.',
    '27168': 'NEW candidate 73211>72222: "Too Wet to Mow" corflute signs x6, mirrors the standing Tennyson decals ruling. Held un-instructed.',
    'DOC0079602': 'F3 candidate 73211>73544: PRINCE2 Agile Foundation, Duen Jaemjamrat. Held un-instructed.',
    '00001400': 'F3 candidate 73211>73544: PSIA Playground Safety Training (Perry, Ford); reversed off by GJ078546, destination unverified. Held un-instructed.',
    '39657': 'F3 candidate 73211>73544: IPWEA Professional Cert AMP; reversed off by GJ078546, destination unverified; $343.20 prior-payment posting trace open. Held un-instructed.',
    'INV-1917': 'CAASie digital-billboard advertising, invoice addressed to Corporate Communications; classification and ownership to confirm.',
    'INV-251010(2)': 'F4: Org Risk WHS Gap Analysis, reversed off by GJ078546 ($11,098.91 leg), destination unverified; consultancy content (73601 home account). Invoice sighted under NA73601.',
    'INV-13157': 'F4: Bendelta Phase 1 Strategic Workshop, reversed off by GJ078546, destination unverified; consultancy content. Invoice sighted under NA73544 (L17).',
}
TELSTRA_NOTE = ('Telstra account 493 5987 935 (PA - Parks); bill sighted 11-Jun, /1.1 exact. GSTE-28 pattern '
                '(AP books incl/1.1 vs disclosed GST, cents-level both directions). Classification query: '
                'branch telephone account arguably telecommunications, not Minor Contracts. Held un-instructed.')
SC_NOTE = ('F5 CLOSED as a trace: Worssell sales credit sighted (17kg old plaques scrapped). NEW candidate: scrap income '
           'to 64411 mirroring the 72111 ruling, plus the cemeteries attribution (plaques are svc 20451; credit sits on Depots PK000001). Held un-instructed.')
TE4838_NOTE = ('F7: cover slip codes PK000055-73212, line posted 1-20241-73211; content is bulk customer SMS (Sinch '
               'INV05480153 sighted with cover slip), arguably neither account. Held for ruling.')

def class_73211(ref, det, amt, wo):
    det = str(det)
    note_extra = ''
    # L3: canonical sighted set
    key = ('GJ078960', '14566') if (ref == 'GJ078960' and '14566' in det) else ref
    sighted = key in SIGHTED_REFS
    l1 = 'A' if ref in L1_AMBER else 'G'
    l3 = 'G' if sighted else 'P'
    l4 = 'P'
    if sighted:
        l4 = 'G' if ref in ('LCC31326', '39657') else 'A'
    note = ''
    l2 = 'G' if sighted else 'P'
    if ref in L2_AMBER_NOTE:
        l2 = 'A'; note = L2_AMBER_NOTE[ref]
    elif ref.startswith('49359879'):
        l2 = 'A'
        note = TELSTRA_NOTE if sighted else 'Telstra Dec-25 bill GAP ($727.95 / $800.75); telecoms classification query rides with the sighted series.'
        if ref == '49359879351225':
            l2, l3, l4 = 'A', 'P', 'P'
    elif ref == 'SC10918':
        l2 = 'A'; note = SC_NOTE
    elif ref == 'TE004838':
        l2 = 'A'; note = TE4838_NOTE
    elif ref == 'GJ079190':
        l2 = 'G'
        note = ('F1: duplicates GJ079210 onto PK000022 ($11,270 for the $5,635 INV-5888); phantom -$5,635 credit on '
                'WO PK000511. Gihani 6-May-2026 instruction sighted (510>PK000023, 511>PK000022): the 510 batch is misposted '
                'or a straight duplicate. Reversal candidate held; narration "Timber Treatment (PK000510)" describes neither leg.')
    elif ref == 'GJ079210':
        l2 = 'G'
        note = 'F1 context: GJ079210 STANDS (clears INV-5888 off 20261/PK000035 to PK000022, direction correct); the duplicate is GJ079190.'
    elif ref == 'GJ078546':
        l2 = 'A'
        if amt < -8000:
            note = 'F4: reverses $11,098.91 off 20001 (WHS Gap Analysis $2,132.63 + Strategic Workshop $8,966.28, sum exact); destination outside this export, 73600-series likely; to confirm.'
        elif amt < 0:
            note = 'F3: reverses PSIA $4,300 + IPWEA $2,788 + chair $370.05 at exactly -$7,458.05; chair is 72313 per the equipment rule, training items 73544 candidates; destinations unverified. Narration "Consulting Pty Ltd" describes none of the three.'
        else:
            note = 'F4 inbound query: Provac INV-00037183 reversal (+$2,050, grave dig in rock with dock hire); plant hire on Minor Contracts; 73123 or a plant hire account are the alternatives.'
    elif ref in ('GJ078945', 'GJ079117'):
        l2 = 'G'; l3 = 'A'
        note = 'INV-0352 split pair (GJ078945/GJ079117): nets $0.00, direction correct; $15,000 moved 20591>20641 15-Apr-2026 per the Document Reconstruction (filed Related/). Underlying spend carries the F2 flag on the invoice line.'
    elif ref == 'GJ078960':
        if '14566' in det:
            note = 'EWN invoice 14566 sighted ($75.00 / $82.50 exact, planned-burns notifications Aug-2025).'
        elif '14768' in det:
            note = 'EWN INV-14768 GAP.'
        else:
            note = 'Ausecology ' + ('INV-4029' if amt > 7000 else 'INV-4030') + ' GAP (via GJ078960).'
    elif ref in ('3223657', '3223717'):
        l3 = 'A'; l4 = 'A'; l2 = 'G'
        note = ('T2 invoice sighted 11-Jun (job 2221939 Parks Depot; Order 802502 reallocated to PO 714635 per the Aisha Wilson '
                'email, filed Related/) - outside the canonical evidence count ($157,051.42); register wins, limb held A pending manifest inclusion.')
    elif ref == 'INV-5888':
        note = 'Sighted 11-Jun, exact (PO 713687 Play Force asset capture - confirms the F1 quantum; content is asset capture, not timber treatment). Original posting on WO PK000511 (work-order pseudo-PK per the Gihani email).'
    elif ref == 'RL195364':
        if abs(amt - 0.01) < 0.005:
            l2 = 'G'; note = 'F6 CLOSED: the $0.01 variance line is Empire\'s GST rounding cent (GST $37.01 on $370.05).'
        else:
            l2 = 'A'; note = 'F3 candidate 73211>72313: Jemma chair (Empire RL195364 sighted, exact; desk/chair components are 72313). Held un-instructed.'
    elif sighted:
        note = 'Sighted 11-Jun-2026 (intake C00228240), reconciles exact on the x1.1 test; ABN checksum PASS; ABR sighting queued.'
    elif ref == '2907368':
        note = 'Three-line set (-$5,100 / $0 / +$5,100) nets $0.00 within the account; reversal and repost; backup unsighted.'
    else:
        note = 'Unsighted; gap row in NA73211_Evidence_Manifest.csv. Attachment flag present but flags are not sighting.'
    if sighted and ref in VENDOR and VENDOR[ref] == 'Lockwise Locksmiths':
        note = 'Lockwise standing order PO 709747; invoice sighted 11-Jun, reconciles exact; ABR sighting queued.'
        if abs(amt) < 0.005:
            note += ' Zero-value companion line.'
    if ref == 'LCC31326':
        l4 = 'G'
    ovr = 'A' if 'A' in (l1, l2, l3, l4) else ('P' if 'P' in (l2, l3, l4) else 'G')
    return l1, '', l2, l3, l4, ovr, note

# ---------------- small-tab row factories ----------------
def rows_73211():
    df = load(EXP + '2026-06-11T100221.018.xlsx')
    out = []
    for _, r in df.iterrows():
        acct = str(r['Account']); svc = acct.split('-')[1]
        pk, section = svc_info(svc)
        ref = str(r['Reference']); amt = float(r['amt'])
        l1, l1e, l2, l3, l4, ovr, note = class_73211(ref, r['Details'], amt, str(r['Work Order']))
        card = 'Craig Logan' if ref == 'TE004838' else ''
        out.append(['73211', svc, pk, section, int(r['Period']), str(r['Date'])[:10], ref,
                    VENDOR.get(ref, ''), card, str(r['Details'])[:58], round(amt, 2),
                    round(amt * 1.1, 2), 'Y' if str(r['Has Attachment']).upper() == 'Y' else 'N',
                    l1, l1e, l2, l3, l4, ovr, note, '', ''])
    return out

def manual_rows(na, specs):
    out = []
    for (svc, per, date, ref, vendor, card, det, amt, incl, att,
         l1, l1e, l2, l3, l4, ovr, note) in specs:
        pk, section = svc_info(svc)
        out.append([na, svc, pk, section, per, date, ref, vendor, card, det,
                    amt, incl, att, l1, l1e, l2, l3, l4, ovr, note, '', ''])
    return out

R73541 = manual_rows('73541', [
    ('20151', 1, '2025-07-08', 'GJ075544', '', '', 'Hannah McCone Butcher & Scott Cameron Ph-88 Kidston', -44.89, -49.38, 'Y',
     'G', '', 'G', 'G', 'G', 'G', 'EOY 2024/25 AP reversal pair, nets $0.00 with the 2/07/2025 line. Kidston St Canungra reads as accommodation, never 73541 content - FY25 advisory, noted for the FY25 file.'),
    ('20151', 10, '2026-04-14', 'GJ078933', '', '', 'TE004218 City of Gold Coast Parking Jul--Parking not', -6.05, -6.66, 'Y',
     'G', '', 'G', 'G', 'G', 'G', 'Recode 73541>73531 posted in-ledger 14-Apr-2026; pair nets $0.00, direction correct. Closes the restored 3-Jun travel flag on TE004218.'),
    ('20151', 1, '2025-06-16', 'TE004218', 'City of Gold Coast', 'Craig Logan', 'CityofGoldCoastParking-General Expenses-Purchase Car', 6.05, 6.66, 'Y',
     'G', 'G', 'G', 'G', 'G', 'G', 'Parking miscoded to conferences; corrected in-ledger by GJ078933 per the travel rule (73531 local SEQ). Pair CLOSED.'),
    ('20641', 6, '2025-12-01', 'TE004922', 'Trailbuilders.org', '', 'INTNL TRANSACTION FEE-Conference Registration-Purcha', 15.66, 17.23, 'Y',
     'G', '', 'G', 'A', 'A', 'A', 'International transaction fee riding on the registration: strictly a financing cost, not a registration fee; materiality does not justify a recode. Advisory only; backup unsighted with the registration.'),
    ('20151', 1, '2025-05-28', '2/07/2025', '', '', 'Hannah McCone Butcher & Scott Cameron Ph-88 Kidston', 44.89, 49.38, 'Y',
     'G', '', 'G', 'G', 'G', 'G', 'EOY 2024/25 AP reversal pair companion line (Creditors invoice); nets $0.00 with GJ075544.'),
    ('20641', 6, '2025-12-01', 'TE004922', 'Trailbuilders.org', '', 'TRAILBUILDERS.ORG-Conference Registration-Purchase C', 626.48, 689.13, 'Y',
     'G', '', 'G', 'A', 'A', 'A', 'Conference registration is correct 73541 content on its face; invoice unsighted - sight the Trailbuilders backup on the PCard reconciliation.'),
])

R7B532 = manual_rows('7B532', [
    ('20451', 3, '2025-09-16', 'IJ073064', 'Bega Road (internal)', '', 'Spoil (Non Fire Ant zone)-Quarry Returns W/E 01-Jan-', 480.00, 480.00, 'Y',
     'G', '', 'G', 'G', '', 'G', 'CLEARED: NA7B532_E02 Bega Road sheet 16-Sep-2025, 30 units x $16 = $480 exact (rate validated). Internal charge, no GST uplift.'),
    ('20451', 1, '2025-07-28', 'IJ072723', 'Bega Road (internal)', '', 'Spoil (Non Fire Ant zone)-Quarry Returns W/E 01-Jan-', 1232.00, 1232.00, 'Y',
     'G', '', 'G', 'G', '', 'G', 'Evidenced as a single charge: NA7B532_E01 sheet 28-Jul-2025, 77 units x $16 = $1,232 exact.'),
    ('20451', 1, '2025-07-28', 'IJ072723', 'Bega Road (internal)', '', 'Spoil (Non Fire Ant zone)-Quarry Returns W/E 01-Jan-', 1232.00, 1232.00, 'Y',
     'G', '', 'G', 'A', '', 'A', 'Probable duplicate over-charge: identical reference, date and narration; the 28-Jul-2025 sheet shows one 77-unit load only. Held AMBER not RED - the corrupted "W/E 01-Jan-1900 Week 0" field keeps a weekly basis in play. Obtain the full week\'s Bega Road sheets.'),
])

R7B214 = manual_rows('7B214', [
    ('20151', 8, '2026-02-17', 'IJ074028', 'CSL Pest Management (internal)', '', 'Parks Depot (Marsden)-181-191 Chambers Flat, Road, M', 82.50, 82.50, 'Y',
     'G', '', 'G', 'G', '', 'G', 'CLEARED on Tier 1: CSL invoice sheet (rodent control half-hour 3-Feb-2026, $82.50, after-hours nil) + Durant email 5-Feb-2026; internal oncharge passed through at invoice price, exact.'),
    ('20392', 7, '2026-01-27', 'IJ073869', 'Transport Operations (internal)', '', 'Tully Memorial Park-Modification to Signage-Transpor', 3600.00, 3600.00, 'Y',
     'G', '', 'A', 'A', '', 'A', 'Transport Operations oncharge backup required. Signage modification reads closer to works than maintenance services; classification confirmed only on sighting.'),
    ('20361', 12, '2026-06-10', 'IJ074845', 'DIV99 (internal)', '', 'Inv# DA/4574-2025/2026 Backflow Annual-Renewal Notic', 3888.00, 3888.00, 'Y',
     'A', '', 'A', 'A', '', 'A', 'DA/4574 required; backflow prevention testing is plausible 7B214 content, the DIV99 renewal-notice framing wants confirmation it is a Parks cost at svc 20361. Posted 10-Jun-2026 (P12, after the 8-Jun SE2 cut).'),
])

R62 = []
for (acct, svc, per, date, ref, det, amt, incl, l1, l2, l3, l4, ovr, note) in [
    ('62121', '20453', 10, '2026-04-10', '505970', 'Burial Interment Fee-E/M01/154-SUSKI', -2812.73, '',
     'G', 'G', 'A', 'A', 'A', 'Burial interment fee SUSKI; invoice 505970 unsighted - evidence limb holds here.'),
    ('62125', '20446', 2, '2025-08-01', '502197', 'C/L01/F/283 - Helen Davies', -2159.00, -2159.00,
     'G', 'G', 'G', 'G', 'G', 'Pre-purchase plot (1 of 3 at $2,159 on invoice 502197, Davies; GST-free, exact). PK000400 does not go to reserve; reversed in full by GJ076150 - recode justified and actioned.'),
    ('62125', '20446', 1, '2025-07-16', '501999', 'Ashes Interments', -1552.73, -1708.00,
     'G', 'G', 'G', 'G', 'G', 'King reissued invoice 501999 (chain closed: 498178 charged GST in error, CN 140775 credited, 501999 GST-correct). Line is the taxable ashes interments $1,552.73 x 1.1 = $1,708.00 exact. Advisory: ashes interment component arguably sits closer to 62121 than 62125.'),
    ('62121', '20453', 11, '2026-05-06', '506368', "HALA'UFIA - B/M01/164", -204.55, '',
     'G', 'G', 'A', 'A', 'A', 'Burial interment fee HALA\'UFIA; invoice 506368 unsighted.'),
    ('62125', '20446', 2, '2025-09-01', 'GJ076150', 'Ashes Interments', 1552.73, 1708.00,
     'A', 'G', 'G', 'G', 'A', 'Garside query actioned: reversal of the 501999 line off PK000400. Destination legs sit outside the supplied export; confirm the receiving PK disburses to svc 20451 (the point of the query).'),
    ('62121', '20131', 1, '2025-07-31', '140799', 'Plot fee charged in error for TIMMINS-B/L06/155', 1878.00, 1878.00,
     'G', 'G', 'G', 'G', 'G', 'CN 140799 (Lohrisch) sighted: Timmins plot-fee error credit against invoice 502146, GST-free, exact match. CLEARED.'),
    ('62125', '20446', 2, '2025-09-01', 'GJ076150', 'C/L01/F/283 - Helen Davies', 2159.00, 2159.00,
     'A', 'G', 'G', 'G', 'A', 'Garside query actioned: reversal of the 502197 line off PK000400. Destination-leg confirmation as above.'),
]:
    pk, section = svc_info(svc)
    R62.append([acct, svc, pk, section, per, date, ref, 'LCC Cemeteries (AR)' if ref not in ('GJ076150',) else '',
                '', det, amt, incl, 'Y', l1, '', l2, l3, l4, ovr, note, '', ''])

TABS = [
    ('73211', 'Minor Contracts (below $20,000)', rows_73211()),
    ('73541', 'Conferences & Seminar Fees', R73541),
    ('7B532', 'Internal - Spoils (Cemeteries)', R7B532),
    ('7B214', 'Internal - Maintenance Services', R7B214),
    ('62121-62125', 'Cemetery Revenue (Garside query)', R62),
]

# checks before writing
t73211 = sum(r[10] for r in TABS[0][2])
assert abs(t73211 - 236378.27) < 0.005, t73211
assert len(TABS[0][2]) == 114
sighted_sum = sum(r[10] for r in TABS[0][2] if r[16] == 'G')
assert abs(sighted_sum - 157051.42) < 0.005, sighted_sum
assert abs(sum(r[10] for r in R73541) - 642.14) < 0.005
assert abs(sum(r[10] for r in R7B532) - 2944.00) < 0.005
assert abs(sum(r[10] for r in R7B214) - 7570.50) < 0.005
assert abs(sum(r[10] for r in R62) - (-1139.28)) < 0.005
print('pre-write reconciliation PASS: 73211 %.2f (sighted %.2f), 73541, 7B532, 7B214, 62121/62125' % (t73211, sighted_sum))

wb = load_workbook('00_Running_Transaction_Listing.xlsx')

def write_tab(name, title, rows):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
    t = ws.cell(1, 1, title)
    t.font = Font(name="Segoe UI", bold=True, size=12, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=NAVY)
    ws.row_dimensions[1].height = 22
    for j, (n, w) in enumerate(COLS, 1):
        c = ws.cell(2, j, n)
        c.font = HDR; c.fill = PatternFill("solid", fgColor=NAVY); c.alignment = ctr; c.border = bd
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[2].height = 28
    r = 3
    for row in rows:
        for j, ((n, w), v) in enumerate(zip(COLS, row), 1):
            if n in RAGCOLS:
                c = ws.cell(r, j, v)
                f, fc = RAG.get(v, ("FFFFFF", "000000"))
                c.fill = PatternFill("solid", fgColor=f)
                c.font = Font(name="Segoe UI", bold=True, color=fc, size=10); c.alignment = ctr
            else:
                c = ws.cell(r, j, v)
                c.font = B
                c.alignment = ctr if n in ("Period", "Date", "Attach", "ExGST", "InclX11", "Svc") else lft
                if n in ("ExGST", "InclX11") and v not in (None, ''):
                    c.number_format = '#,##0.00'
                if r % 2 == 0:
                    c.fill = band
            c.border = bd
        r += 1
    ws.freeze_panes = "G3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLS))}{r-1}"
    return r - 3

def cnt(rows, col, *codes):
    i = [n for n, _ in COLS].index(col)
    return sum(1 for r in rows if r[i] in codes)

for name, label, rows in TABS:
    tot = sum(r[10] for r in rows)
    extra = '; export nets -$1,139.28, review scope $3,711.73 (register)' if name == '62121-62125' else ''
    sub = ' - tax limb n/a (internal charges, no GST uplift)' if name in ('7B532', '7B214') else ''
    n = write_tab(name, f"NA {name.replace('62121-62125','62121/62125')} {label} — line-level review ({len(rows)} lines, ${tot:,.2f} ex-GST{extra}){sub}", rows)
    print(f'tab {name}: {n} lines, {tot:,.2f}')

# ---- Summary rewrite: insert the five rows before the TOTAL row ----
sw = wb['Summary']
# locate TOTAL row
tot_row = next(r for r in range(3, sw.max_row + 1) if str(sw.cell(r, 1).value).startswith('TOTAL'))
sw.insert_rows(tot_row, 5)
labels = {'73211': 'Minor Contracts (below $20,000)', '73541': 'Conferences & Seminar Fees',
          '7B532': 'Internal - Spoils (Cemeteries)', '7B214': 'Internal - Maintenance Services',
          '62121-62125': 'Cemetery Revenue (Garside query)'}
r = tot_row
for name, label, rows in TABS:
    tot = round(sum(x[10] for x in rows), 2)
    rev = sum(1 for x in rows if x[18] != 'P')
    vals = [f"{name.replace('62121-62125','62121/62125')} {labels[name]}", len(rows), tot,
            f"{rev/len(rows)*100:.0f}%",
            f"{cnt(rows,'L1 Svc/PK','R')} / {cnt(rows,'L1 Svc/PK','A')}",
            f"{cnt(rows,'L2 NA','R')} / {cnt(rows,'L2 NA','A')}",
            f"{cnt(rows,'L3 Evid','R')} / {cnt(rows,'L3 Evid','A')}",
            f"{cnt(rows,'L4 Tax','R')} / {cnt(rows,'L4 Tax','A')}",
            f"{cnt(rows,'Overall','R')} / {cnt(rows,'Overall','A')} / {cnt(rows,'Overall','P')}"]
    if name == '62121-62125':
        vals[8] += '  (revenue; export nets -$1,139.28)'
    for j, v in enumerate(vals, 1):
        c = sw.cell(r, j, v)
        c.font = B; c.alignment = lft if j == 1 else ctr; c.border = bd
        if j == 3:
            c.number_format = '#,##0.00'
    r += 1
# TOTAL row: live SUMs over the account rows (rows 4 .. r-1)
sw.cell(r, 1, 'TOTAL line-level').font = Font(name="Segoe UI", bold=True)
sw.cell(r, 2, f'=SUM(B4:B{r-1})').font = Font(name="Segoe UI", bold=True)
c = sw.cell(r, 3, f'=SUM(C4:C{r-1})')
c.font = Font(name="Segoe UI", bold=True); c.number_format = '#,##0.00'
for j in range(1, 10):
    sw.cell(r, j).border = bd

wb.save('00_Running_Transaction_Listing.xlsx')
print('listing saved')
