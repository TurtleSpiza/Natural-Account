#!/usr/bin/env python3
"""Extend the Tracker and the Status workbook for the five 11-Jun handover
workstreams. SE2-coverage convention holds on the 87-account sheets (the 73511
precedent); register-basis differences are recorded, never papered over, and a
live tie-out to the register total $502,932.11 is added to the Summary."""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule

D = '11-Jun-2026'

# ---------------- Tracker ----------------
wb = load_workbook('00_Parks_4090000_NAReview_Tracker.xlsx')
ws = wb['Tracker']
BASE = Font(name='Segoe UI', size=10)

def trow(account):
    for r in range(8, 95):
        if str(ws.cell(r, 2).value) == account:
            return r
    raise KeyError(account)

updates = {
    '73211': ('Closed', 236077.27, 0,
              'Reviewed in full 11-Jun-2026: 114 lines reconcile to the 11-Jun SE2 to the cent at $236,378.27 (register basis); '
              'this row carries the 8-Jun SE2 actual $236,077.27, the $301.00 difference being the P12 Lockwise line 2908883 '
              'posted 8-Jun-2026 after the SE2 cut. Intake C00228240: 43 documents sighted, $157,051.42 (66.4%) evidenced. '
              'Findings F1-F7 held for ruling; no recode staged.'),
    '7B532': ('Partial', 2944.00, 3088.00,
              'Three internal journal lines on 1-20451 reviewed 11-Jun-2026 ($2,944.00); $16/unit rate validated from the Bega '
              'Road sheets; probable $1,232 IJ072723 duplicate held AMBER pending the full week\'s sheets. Remainder untested. '
              'Tax limb n/a (internal).'),
    '7B214': ('Closed', 3682.50, 0,
              'Full 3-line internal population reviewed 11-Jun-2026; register carries $7,570.50 including the P12 IJ074845 '
              '$3,888.00 posted 10-Jun-2026 after the 8-Jun SE2 cut (basis difference $3,888.00). IJ074028 $82.50 cleared on '
              'Tier 1; $7,488.00 oncharge backups unsighted. Tax limb n/a (internal).'),
    '73541': ('Closed', 642.14, 0,
              'Reviewed in full 11-Jun-2026: 6 lines net $642.14. TE004218 $6.05 parking miscode CLOSED (recoded to 73531 '
              'in-ledger via GJ078933, pair nets $0.00, direction correct); GJ075544 EOY pair nets out (FY25 Canungra advisory). '
              'Open: Trailbuilders TE004922 backup unsighted ($642.14 combined).'),
}
for acct, (status, rev, unt, note) in updates.items():
    r = trow(acct)
    for col, v, fmt in ((9, status, None), (10, rev, '#,##0.00'), (11, unt, '#,##0.00'), (12, note, None)):
        c = ws.cell(r, col, v)
        c.font = BASE
        if fmt:
            c.number_format = fmt
        if col == 12:
            c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    print(f'tracker {acct} row {r}: {status}, reviewed {rev}')

# supplementary row below TOTAL (row 95), outside all KPI ranges (8:94)
ws.cell(96, 3, 'Supplementary - outside the 87-account expense scope').font = Font(name='Segoe UI', bold=True, italic=True, size=9, color='595959')
vals = ['SUPP', '62121/62125', 'Cemetery Revenue (Garside query)', None, None, None, None, 'revenue', 'Closed*', 3711.73, None,
        'Garside query reviewed 11-Jun-2026 (register row, $3,711.73): GJ076150 already reversed both PK000400 pre-purchase '
        'lines in-ledger; King five-document chain closed, nil GST errors. Residue: destination-leg confirmation (svc 20451) '
        'and burial invoices 505970/506368 unsighted. *Excluded from every KPI above (revenue account, not in the 87).']
for j, v in enumerate(vals, 1):
    if v is None:
        continue
    c = ws.cell(97, j, v)
    c.font = BASE
    if j == 10:
        c.number_format = '#,##0.00'
    if j == 12:
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
wb.save('00_Parks_4090000_NAReview_Tracker.xlsx')
print('tracker saved')

# ---------------- Status workbook ----------------
wb = load_workbook('00_NA_Status_All_Accounts_11-Jun-2026.xlsx')
aa = wb['All Accounts']

def arow(account):
    for r in range(5, 92):
        if str(aa.cell(r, 2).value) == account:
            return r
    raise KeyError(account)

ROWS = {
 '73211': dict(status='Fully reviewed', rev=236077.27, unt=0,
  rags=('AMBER', 'AMBER', 'AMBER', 'AMBER', 'AMBER'),
  find=('Reconciles to the 11-Jun SE2 to the cent: $236,378.27 register basis, 114 lines, 23 cost codes (this sheet carries '
        'the 8-Jun SE2 actual $236,077.27; the $301.00 difference is the P12 Lockwise line 2908883 posted 8-Jun-2026). '
        'Intake C00228240 applied 11-Jun: 43 documents sighted, $157,051.42 (66.4%) evidenced, every document exact on the '
        'x1.1 test, 18 ABN checksums PASS. Findings: F1 $5,635 GJ079190 duplicates GJ079210 onto PK000022 (Gihani 6-May-2026 '
        'instruction sighted: 510>PK000023, 511>PK000022); F2 ~$98,005 above-threshold candidates to 73212 (LCC31326 '
        '$25,848.08, Destination Trails $43,114, Chronicle $29,042.45); F3 training/chair items (PRINCE2 $2,440, PSIA $4,300, '
        'IPWEA $2,788, chair $370.05); F4 GJ078546 destinations unverified plus Provac $2,050 inbound query; F7 TE004838 '
        'cover slip vs posting mismatch; F5 and F6 CLOSED. Six new classification candidates (Firesight 73601, Chronicle '
        '73566/survey, Tennyson 72222, SC10918 64411, Telstra telecoms, INCON materials). GSTE-28 Telstra cents-level pattern.'),
  rem=('1 Spero rulings on F1 (reversal and re-post destination), the F2/F3 candidate sets and the six new candidates '
       '(single sweep). 2 Sight the remaining gap documents ($79,326.85 incl 2906937 $10,200, 2908705 $17,000, TBM series '
       '$15,604.11, Ausecology $14,840). 3 Run the ABR sightings queue (18 suppliers; INCON and IPWEA done). 4 No recode '
       'dollar populates without the ruling.'),
  evid='02_NA73211_Minor_Contracts/ (Verification Record; 43 evidence PDFs + manifest with GAP rows; Related/)',
  jrnl='n/a (no recode staged; candidates held un-instructed)'),
 '7B532': dict(status='Partial', rev=2944.00, unt='=D{r}-H{r}',
  rags=('GREEN', 'GREEN', 'AMBER', '', 'AMBER'),
  find=('$2,944.00 over 3 internal journal lines on 1-20451-7B532 PK000403 (Cemeteries spoils). $16 per unit rate validated '
        '(IJ073064 $480 = 30 units x $16 exact; 28-Jul-2025 sheet 77 units x $16 = $1,232 exact). IJ073064 cleared GREEN; one '
        'IJ072723 $1,232 posting evidenced; the identical second posting is a probable duplicate over-charge, held AMBER not '
        'RED because the corrupted week-ending narration ("W/E 01-Jan-1900 Week 0") keeps a weekly charge basis in play. Tax '
        'limb n/a (internal charges, no GST uplift). Untested remainder sits outside the 3-line export.'),
  rem=('1 Obtain the Bega Road sheets for the week 28-Jul-2025 to 3-Aug-2025 to resolve the probable $1,232 duplicate. '
       '2 Raise the quarry-returns week-ending field corruption with the system owner (affects all 7B532 narrations).'),
  evid='04_NA7B532_Internal_Spoils/ (Verification Record; evidence manifest; two Bega Road sheets)',
  jrnl='n/a (no recode staged)'),
 '7B214': dict(status='Fully reviewed', rev=3682.50, unt=0,
  rags=('AMBER', 'AMBER', 'AMBER', '', 'AMBER'),
  find=('Full 3-line internal population reviewed; the register carries $7,570.50 including the P12 IJ074845 $3,888.00 posted '
        '10-Jun-2026 after the 8-Jun SE2 cut (this sheet\'s accum actual $3,682.50; basis difference $3,888.00). IJ074028 '
        'Marsden rodent control $82.50 cleared on Tier 1 (CSL invoice sheet + Durant email 5-Feb-2026; oncharge passed through '
        'at invoice price, exact). Remaining $7,488.00 unsighted: IJ073869 Tully Memorial Park signage modification $3,600.00 '
        '(Transport Operations oncharge; reads closer to works than maintenance services) and IJ074845 DA/4574 backflow annual '
        'renewal notices $3,888.00 (svc 20361 attribution to confirm). Tax limb n/a (internal charges).'),
  rem=('1 Obtain the Transport Operations oncharge backup for IJ073869 ($3,600.00). 2 Obtain DA/4574 for IJ074845 ($3,888.00) '
       'and confirm the svc 20361 attribution.'),
  evid='05_NA7B214_Internal_Maintenance/ (Verification Record; evidence manifest; CSL sheet + Durant email)',
  jrnl='n/a (no recode staged)'),
 '73541': dict(status='Fully reviewed', rev=642.14, unt=0,
  rags=('GREEN', 'GREEN', 'AMBER', 'AMBER', 'AMBER'),
  find=('6 lines net $642.14 (svc 20641 PCard plus closed recode and reversal pairs). TE004218 $6.05 Gold Coast parking '
        'miscode CLOSED: recoded to 73531 via GJ078933 (14-Apr-2026), pair nets $0.00, direction correct, closing the restored '
        '3-Jun travel flag. GJ075544 +/-$44.89 EOY 2024/25 AP reversal pair nets out (Canungra accommodation content is an '
        'FY25 advisory; accommodation never belongs on 73541). Open: TE004922 Trailbuilders conference registration $626.48 '
        'unsighted (correct 73541 content on its face) plus a $15.66 international transaction fee, strictly a financing cost, '
        'immaterial, advisory only.'),
  rem='1 Sight the Trailbuilders registration backup for TE004922 ($642.14 combined) on the PCard reconciliation.',
  evid='06_NA73541_Conferences/NA73541_Notes_and_Kinatico_Intake.md (none sighted)',
  jrnl='n/a (TE004218 recode already in-ledger via GJ078933)'),
}
BASE9 = Font(name='Segoe UI', size=9)
wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)
for acct, d in ROWS.items():
    r = arow(acct)
    aa.cell(r, 7, d['status']).font = BASE9
    c = aa.cell(r, 8, d['rev']); c.font = BASE9; c.number_format = '#,##0.00'
    unt = d['unt']
    c = aa.cell(r, 9, unt.format(r=r) if isinstance(unt, str) else unt)
    c.font = BASE9; c.number_format = '#,##0.00'
    for j, v in zip(range(10, 15), d['rags']):
        c = aa.cell(r, j, v if v else None)
        c.font = BASE9
        c.alignment = Alignment(horizontal='center', vertical='center')
    for col, key in ((15, 'find'), (16, 'rem'), (17, 'evid'), (18, 'jrnl')):
        c = aa.cell(r, col, d[key]); c.font = BASE9; c.alignment = wrap
    aa.cell(r, 19, D).font = BASE9
    aa.row_dimensions[r].height = 110
    print(f'status {acct} row {r}: {d["status"]}')

# supplementary block rows 93-94 (outside the 5:91 formula and CF ranges)
lab = aa.cell(93, 1, 'SUPPLEMENTARY - outside the 87-account expense scope (revenue; carried on the register, excluded from the KPI ranges above)')
lab.font = Font(name='Segoe UI', bold=True, italic=True, size=9, color='595959')
supp = ['SUPP', '62121/62125', 'Cemetery Revenue (Garside query)', None, None, 'revenue', 'Fully reviewed', 3711.73, 0,
        'AMBER', 'GREEN', 'AMBER', 'AMBER', 'AMBER',
        ('Garside query (29-Aug-2025) already actioned in-ledger: GJ076150 (1-Sep-2025) reversed both PK000400 pre-purchase '
         'lines ($1,552.73 + $2,159.00 = $3,711.73) off 1-20446-62125. Five sighted documents reconcile to the cent; the King '
         'paper trail is closed (498178 charged GST in error, CN 140775 credited, 501999 reissued GST-correct) and CN 140799 '
         'Timmins plot-fee error cleared. GST treatment consistent (plots GST-free, interments taxable); nil GST Errors '
         'Register entries. $ incl GST is per-line, not a blanket 1.1 uplift.'),
        ('1 Pull the GJ076150 destination legs and confirm the receiving PK disburses to svc 20451. 2 Sight burial-fee '
         'invoices 505970 ($2,812.73 SUSKI) and 506368 ($204.55 HALA\'UFIA). 3 Advisory: the ashes interment component '
         'arguably sits closer to 62121 than 62125. 4 Scope call: the 56-line event-fee export (-$61,364.17) awaits a ruling.'),
        '03_Cemetery_Revenue_62121_62125/ (Verification Record; 5 evidence docs + Garside email)',
        'n/a (no recode staged; GJ076150 already in-ledger)', D]
for j, v in enumerate(supp, 1):
    if v is None:
        continue
    c = aa.cell(94, j, v)
    c.font = BASE9
    if j in (8, 9):
        c.number_format = '#,##0.00'
    if j in (15, 16, 17, 18):
        c.alignment = wrap
    if j in range(10, 15):
        c.alignment = Alignment(horizontal='center', vertical='center')
aa.row_dimensions[94].height = 95
# RAG + status conditional formatting for the supplementary row
for word, fill, fcol in (('GREEN', '375623', 'FFFFFF'), ('AMBER', 'FFF2CC', '000000'), ('RED', 'C00000', 'FFFFFF')):
    aa.conditional_formatting.add('J94:N94', CellIsRule(
        operator='equal', formula=[f'"{word}"'],
        fill=PatternFill('solid', fgColor=fill), font=Font(color=fcol, size=9, name='Segoe UI', bold=True)))

# ---------------- Summary tab: headlines, basis note, live register tie-out ----------------
sm = wb['Summary']
heads = [
 '  - Derived workbooks extended 11-Jun-2026 (fourth pass): the five handover workstreams (73211, 62121/62125, 7B532, 7B214, 73541) now carry rows here, in the Tracker and in the Running Transaction Listing (133 new listing lines).',
 '  - 73211 worked and evidenced: intake C00228240 applied, 43 documents sighted, $157,051.42 of $236,378.27 (66.4%), all exact on the x1.1 test, 18 ABNs PASS; F5/F6 closed, F7 raised; GSTE-28 Telstra pattern recorded.',
 '  - 73211 candidates held un-instructed (no recode dollar populated): F1 GJ079190 reversal, F2 ~$98,005 to 73212, F3 training/chair items, six new classification candidates.',
 '  - 7B532 probable $1,232 IJ072723 duplicate held AMBER; 7B214 $7,488 oncharge backups outstanding; 62121/62125 destination-leg confirmation is the Garside residue; 73541 closes on the Trailbuilders backup.',
 '  - Hard gates NONE. Register 21 rows, $502,932.11 ex-GST reviewed (recalc-verified); 73123 remains OPENED, not reconciled (gap $50,338.64) and blocks its line work.']
for i, h in enumerate(heads):
    sm.cell(16 + i, 2, h).font = Font(name='Segoe UI', size=10)
sm.cell(21, 2, ('Basis note: SE2-coverage Reviewed-$ differs from the register reviewed total of $502,932.11 by $9,026.23: '
                'measurement-basis differences on 73511 ($1,117.08, FBT batch pairs) and 73512 ($8.42) carried from the '
                '10-Jun merge, post-SE2 P12 postings on 73211 ($301.00, Lockwise 2908883 posted 8-Jun-2026) and 7B214 '
                '($3,888.00, IJ074845 posted 10-Jun-2026), and the 62121/62125 revenue supplementary ($3,711.73) outside the '
                '87-account expense scope. The tie-out below recomputes the bridge live.')).font = Font(name='Segoe UI', size=10, italic=True)

t = sm.cell(23, 2, 'Register tie-out (live): SE2-coverage Reviewed-$ bridged to the Account Review Register')
t.font = Font(name='Segoe UI', bold=True, size=11, color='1F3864')
rows = [
 ('Reviewed $ (SE2-coverage basis, 87 accounts)', "=SUM('All Accounts'!H5:H91)", ''),
 ('+ 73511 basis difference (register $5,384.57 net of FBT batch pairs)', "=5384.57-SUMIF('All Accounts'!B5:B91,\"73511\",'All Accounts'!H5:H91)", ''),
 ('+ 73512 basis difference (register $5,520.97 post-SE2)', "=5520.97-SUMIF('All Accounts'!B5:B91,\"73512\",'All Accounts'!H5:H91)", ''),
 ('+ 73211 basis difference (register $236,378.27; P12 Lockwise 2908883)', "=236378.27-SUMIF('All Accounts'!B5:B91,\"73211\",'All Accounts'!H5:H91)", ''),
 ('+ 7B214 basis difference (register $7,570.50; P12 IJ074845)', "=7570.5-SUMIF('All Accounts'!B5:B91,\"7B214\",'All Accounts'!H5:H91)", ''),
 ('+ 62121/62125 supplementary (revenue, outside the 87)', "='All Accounts'!H94", ''),
 ('= Register reviewed total (computed)', '=SUM(E24:E29)', 'bold'),
 ('Register total per 00_Account_Review_Register.xlsx, 21 rows, 11-Jun-2026', 3711.73, 'input'),
 ('Tie-out', '=IF(ABS(E30-E31)<0.005,"TIES","CONFLICT - investigate before relying on this sheet")', 'check'),
]
rows[7] = ('Register total per 00_Account_Review_Register.xlsx, 21 rows, 11-Jun-2026', 502932.11, 'input')
for i, (label, val, kind) in enumerate(rows):
    r = 24 + i
    sm.cell(r, 2, label).font = Font(name='Segoe UI', size=10, bold=(kind in ('bold', 'check')))
    c = sm.cell(r, 5, val)
    c.number_format = '#,##0.00'
    if kind == 'input':
        c.font = Font(name='Segoe UI', size=10, color='0000FF')  # input constant convention
    elif kind == 'check':
        c.font = Font(name='Segoe UI', size=11, bold=True, color='375623')
        c.number_format = 'General'
    else:
        c.font = Font(name='Segoe UI', size=10, bold=(kind == 'bold'))

# ---------------- Open Items rebuild ----------------
oi = wb['Open Items']
oi.cell(2, 1, ('Mirrors 00_Outstanding_Evidence_and_Actions.md as at 11-Jun-2026 (fourth pass)  |  Documents to sight, Spero '
               'decisions, internal work  |  Hard gates: NONE'))
for r in range(5, oi.max_row + 1):
    for c in range(1, 6):
        oi.cell(r, c).value = None
ITEMS = [
 ('Document', '73511', 'Warnholtz invoice 001 ($744.30, narrated Training)', 'GAP; payee not in employee map'),
 ('Document', '73511', 'Receipts for the PCard/reimbursement net population', 'GAP; per-line GST-free food check pending'),
 ('Document', '73511', 'Greenbank engagement residuals: the $700.00 room hire payment trace and the $200.00 bond return', 'GAP; non-blocking on the INV-9747 line'),
 ('Document', '73533', 'ibis Melbourne (Accor) tax invoice for the $288.18 accommodation', 'Advisory; GST waived by Spero, ITC exposure noted'),
 ('Document', '73140', 'Sample of CV Check AP invoices', 'GAP; numbers unknown'),
 ('Document', '73140', 'FY26 73140 ledger extract, to apply the five held Kinatico invoices (P00075671, P00081314, P00082255, P00083290, P00086890; ABN 25 111 728 842, PO 801790)', 'NARROWED 11-Jun-2026: invoices in hand in 07_Intake_Held_Kinatico/ (Kinatico is the CVCheck parent); held unapplied per the evidence rule, zero AP document hits in the 11-Jun exports'),
 ('Document', '73564', 'SR backups: SR219175 (Robertson), SR-0315517 (Fry)', 'GAP; SR235810 (Dresman) CLOSED 10-Jun; lines journal-ready on directive'),
 ('Document', '73564', 'Other no-attachment SR backups and five external invoices (2412569490, INAU004378, GJ076194, GJ075813, TE005499)', 'GAP; six off-home lines still open'),
 ('Document', '72111 / 72114', 'Full ledger remainders', 'GAP; large untested balances'),
 ('Document', '73128', 'Per-line evidence split set (56 PDFs; all 50 TCB invoices + bond support)', 'Packaging GAP only; all sighted 10-Jun-2026, SHA-256 fingerprints in NA73128_Workstream_Manifest_10-Jun-2026.csv'),
 ('Document', 'n/a (unplaced)', 'BCC King George Square parking receipt 89991, 4-Jun-2025, $42.15 net + $4.21 GST (card ...1627 Anita Moore)', 'Identified-unplaced; FY25 document, no in-scope FY26 match; likely FY25 73531; archived _Sources_10-Jun-2026/ pending instruction'),
 ('Document', 'n/a (unplaced)', 'Six IPWEA documents (R58890, 40310, 40309, two QNT Stripe receipts x3 copies)', 'Identified-unplaced; attendees not in Parks; archived _Sources_10-Jun-2026/ pending instruction'),
 ('Document', '73512', 'Preussner 12052026 receipt, $27.79 (last reimbursement gap)', 'GAP'),
 ('Document', '73512', 'Quote 5468 vendor tax invoice, $1,656.50 (vendor unknown; largest unevidenced line)', 'GAP'),
 ('Document', '73512', 'Salisnew direct claim receipts ~$234 (six refs)', 'GAP'),
 ('Document', '73512', 'Micro-line population PCard docs (refs 1179861-1211218, TE004278-TE005497, ~100 lines)', 'GAP or Spero waiver'),
 ('Document', '73544', 'Healthy Land & Water tax invoice for the BAL workshop $1,767', 'NARROWED 10-Jun-2026: ABR + ASIC dossier sighted; the compliant tax invoice itself remains the gap, tax limb AMBER'),
 ('Document', '73544', 'Internal journal backups x6 (~$10,600: IJ074222, GJ078710, GJ078895, GJ077155, GJ079267, GJ078296)', 'GAP; GJ079650 CLOSED 10-Jun; GJ076597 CLOSED 11-Jun-2026'),
 ('Document', '73544', 'IPWEA $343.20 part-payment posting trace (39657)', 'GAP'),
 ('Document', '73544', 'Preussner GJ078933 -$3,044.65 destination query', 'REOPENED 11-Jun-2026: 73545 versus 7B411 on the $931.20 leg; one ruling closes it'),
 ('Document', '73212', 'Vendor invoices for s4090230 standing-order lines (~$30,000)', 'GAP; GJ076924 backup CLOSED 11-Jun-2026 (PK000415/PK000429 not in maps - documented 11-Jun in lcc-coding-review as unmapped journal PKs pending confirmation)'),
 ('Document', '73212', 'Contract documents for the $20,000 floor test: LCC-09-2022 (Harpley), Aust Care standing order, Play Force engagement', 'GAP; the intake Play Force AP ledger ($97,347.39, INV-7594 to INV-7733) feeds the Play Force floor test'),
 ('Document', '73601', 'Eight Org Risk Consulting invoices (INV-250714, INV-250811 to 250813, INV-250910 to 250913), $43,871.03', 'GAP; INV-251010(2) $2,132.63 sighted and exact'),
 ('Document', '73123', 'Export-vs-SE2 reconciliation, gap $50,338.64 (export $3,951,109.43 vs SE2 $3,900,770.79)', 'GAP; suspected export row duplication and/or P12 timing; BLOCKS all 73123 line work'),
 ('Document', '73211', 'Remaining evidence gaps after intake C00228240: 2906937 ($10,200), 2908705 ($17,000), TBM 00097461/97583/97626/97780 ($15,604.11), Ausecology INV-4029/4030 ($14,840 via GJ078960), Provac INV-00037183 ($2,050), Telstra Dec-25 ($727.95), T2 3222869/3223786 ($1,820), 698933AU/713170AU ($2,433.44), EWN INV-14768, 9016 ($1,710), 00735 ($200), 34828 ($546.68), INV-38877 ($2,250), 00059013 + 00031055/79/85/86 + 00060617/47 series ($4,106.20)', 'NARROWED 11-Jun-2026: 43 documents SIGHTED, $157,051.42 of $236,378.27 (66.4%) evidenced, all exact; gap rows restated in NA73211_Evidence_Manifest.csv'),
 ('Document', '73211', 'IPWEA $343.20 part-payment posting trace', 'NARROWED 11-Jun-2026: invoice 39657 shows Payment Received $343.20 ($312.00 ex + $31.20 GST); the posting location remains the trace'),
 ('Document', 'ALL NEW SUPPLIERS', 'ABR sightings queue: Destination Trails, Firesight, Reactive Generators, Chronicle Rip, Play Force, RST, CAASie, Tennyson, T2, Lockwise, IAS, PSIA, DDLS, EWN, Worssell, Empire, Telstra, Message4U', 'GAP; checksums all PASS; INCON and IPWEA ABR already sighted'),
 ('Document', '7B532', 'Bega Road sheets for the week 28-Jul-2025 to 3-Aug-2025', 'GAP; resolves the probable $1,232 IJ072723 duplicate'),
 ('Document', '7B214', 'Transport Operations oncharge backup for IJ073869 ($3,600); DA/4574 for IJ074845 ($3,888)', 'GAP; $7,488 of $7,570.50 unsighted'),
 ('Document', '62121/62125', 'GJ076150 destination legs (confirm receiving PK disburses to svc 20451); invoices 505970 ($2,812.73 SUSKI) and 506368 ($204.55 HALA\'UFIA)', 'GAP; the destination confirmation is the residue of the Garside query'),
 ('Document', '73541', 'Trailbuilders registration backup for TE004922 ($626.48 + $15.66 fee)', 'GAP; PCard reconciliation'),
 ('Decision', '73511', 'Ruling on the candidate set: $762.33 to 73512 (amenity $334.71, non-employee $301.71, training catering $125.91), $744.30 Warnholtz to 73544, $23.56 utensils to 72111; Daily Blooms $88.14 customer-gift class', 'Held'),
 ('Decision', '72312', '72312 to 72313 full-reversal - journal-ready (PK000493 Flying Gang sub-PK ADDED to lcc-coding-review 11-Jun-2026; pk_validator now passes)', 'Journal-ready'),
 ('Decision', '73564', '73564 to 73566 software recode and the Body2 move-or-stay; the nine off-home PK lines (three confirmed and staged, six open)', 'Held / staged'),
 ('Decision', '73563', 'WINC systemic recode out of 73563; TE005250 riser destination (72313); 73512 moves are Reason B (50/50 batch data check)', 'Journal-ready'),
 ('Decision', '72111', 'PPE to 72113; STIHL to 72315; Posca to 73563; scrap income to 64411; desk-riser 72313 advisory. Tennyson 72222 CLOSED 10-Jun (Doc46 J22)', 'Journal-ready'),
 ('Decision', '72114', 'Extending the Reali PK split to Mar-2026 (Namoa = Cemeteries PK000439, cross-section); net-movement vs full-reversal format; stores-issue footwear/respiratory-PPE class candidates (single ruling)', 'Held'),
 ('Decision', '73215', 'Full account to 73126 Landscapers & Gardeners, $4,739.00, full-reversal, 3 pairs, nets $0.00', 'Candidate HELD un-instructed (Tier 1 complete on 100% of dollars; source leg trips the 73215 Alliance restriction flag by design)'),
 ('Decision', '73128', 'Full milk spend to 73512 - instructed by Spero 10-Jun-2026, journal-ready (122 lines, 61 pairs, nets $0.00, direction verified). Remaining: $61.11 BAS adjustment (GSTE-07 to 22) and per-line evidence packaging', 'Instructed'),
 ('Decision', '73433', 'QUT parking 73433 to 73531 $34.55 - instructed 10-Jun-2026, in the live journal', 'Instructed'),
 ('Decision', '73544', 'PK-split recode INSTRUCTED 11-Jun-2026 and staged ($2,131.46: GJ079650 chargebacks $949.64 + ACCA split $1,181.82, 9 lines, nets $0.00, re-verified)', 'Instructed'),
 ('Decision', '73544', 'Still held: Bendelta $8,966.28 to 73601; BAL workshop $1,767 (73541 vs 73544, tied to the HLW invoice); Warnholtz $744.30; Barter blue card 73433 candidate; Pryor $214.00; Hambly $294.42; Preussner destination query', 'Held'),
 ('Decision', '73212', 'A1 Strategy $15,100 to 73601 (73603 alternative) - Tier 1 complete; Aust Care PK $630 PK000469 to PK000496', 'Candidates held'),
 ('Decision', '73512', 'Micro-line evidence waiver (~100 FBT-batch transfer lines: PCard docs or waiver)', 'Held'),
 ('Decision', 'BAS / GST', 'GSTE-25 ($11.73 Blue Dog) and GSTE-26 ($9.52 Barter) ITC reversals - no GENJNL movement', 'Spero instruction pending'),
 ('Decision', '73211', 'F1 GJ079190 ruling, ENRICHED 11-Jun: Gihani 6-May-2026 email directs 510>PK000023, 511>PK000022; both journals posted to PK000022. Ruling covers (a) full reversal and (b) whether the 510 batch re-posts to PK000023. GJ079210 stands; INV-5888 quantum invoice-confirmed', 'Held'),
 ('Decision', '73211', 'F3 training and chair recodes - source evidence COMPLETE (PSIA, IPWEA, DDLS, Empire all sighted): PRINCE2 $2,440 to 73544; PSIA $4,300 + IPWEA $2,788 to 73544; chair $370.05 to 72313', 'Destination ruling only'),
 ('Decision', '73211', 'New candidates from intake C00228240 (single ruling sweep): Firesight $12,800 to 73601; Chronicle INV-0868 $5,812.45 to 73566; Chronicle mapping/imagery $23,230 survey-services class; Tennyson 27168 $1,058 to 72222; SC10918 -$154 to 64411 + cemeteries attribution; TE004838 $242.40 (F7 mismatch); Telstra series $9,237.48 telecoms class; INCON $25,848.08 materials-not-contracts', 'Held'),
 ('Decision', '73211', 'F2 73211 to 73212 recodes ~$98,005 (LCC31326, Destination Trails series, Chronicle series) - all three series invoice-evidenced in full; treatment ruling pending; total contract value confirmation from contract files or standing orders', 'Held; no recode dollar populates without the ruling'),
 ('Decision', 'Scope', 'Unreviewed queue (parsed and totalled only): 72112 $27,073.10, 72113 $25,244.22, 7B111 $34,840.97, 7B114 $204,911.46, 7B213 $3,996.80, 7B411 $13,130.00, 7BZ11 $10,929.72, 72311 $159.42; plus the 62121 event-fee export (56 lines, -$61,364.17)', 'Confirm whether each joins the review'),
 ('Internal', 'Travel', 'Restored 3-Jun travel flags: Cabcharge Jan-2026 statement $787.74 per-traveller split (mixed 73531/73533). TE004218 CLOSED 11-Jun (GJ078933 verified)', 'Open (Cabcharge only)'),
 ('Internal', '72114', 'Work the Reali set at full scope (Mar-2026 officer detail in hand; GJ079517 $45.29 cross-reference); review the 72111 and 72114 remainders', 'Open'),
 ('Internal', '73511 / 73512', 'Identify Hoffmann, Bahry, Warnholtz and Salisnew (payees not in employee map); confirm Mann and Rosmarin sections; classify four unannotated 73511 lines ($97.27)', 'Open'),
 ('Internal', '73544', "Confirm Burke's Play Safety recode basis: home service resolved 10-Jun as 20362 / PK000435; apply when the recode is instructed", 'Open'),
 ('Internal', '72114', "Item-level tie-out of the GJ078546 redistribution to the five Reali invoices' per-officer detail (dedicated pass)", 'Open'),
 ('Internal', 'Scripts', 'build_listing.py does not rebuild the 73511 tab nor the five workstream tabs added 11-Jun-2026 (73211, 73541, 7B532, 7B214, 62121/62125 - built directly); build_master.py LINE_NAS extended 11-Jun for the four expense tabs', 'Queued'),
 ('Internal', 'People', 'Service confirmations (one EOM People and Positions pass): Anita Moore home service; TE005281 cardholder (QUT); Jennifer Moir; Leighton Funk', 'Open'),
 ('Internal', 'Scripts', 'build_listing.py and build_master.py do not yet cover the four merged accounts (73433, 73512, 73544, 73212) - TechOne exports not packaged; obtain exports, build tabs', 'Queued'),
 ('Internal', '73211', 'SC10918 plaque credit attribution (-$154 on Depots PK000001; plaques are svc 20451) - original purchase line not in the 73211 listing; trail starts with AP', 'Open'),
 ('Internal', '7B532', 'Quarry returns week-ending field corruption ("W/E 01-Jan-1900 Week 0") on all 7B532 narrations - raise with the system owner', 'Open'),
 ('Internal', 'Journals', 'GJ078546 / GJ079190 / GJ079210 narration quality (narrations describe neither leg) - note for the journal preparer', 'Open'),
 ('Register decision', 'Scope', 'Untested remainders: 72111 ($63,621.56) and 72114 ($25,176.66); 73564 untested remainder closed', 'Open'),
 ('Register decision', 'Sequence', 'Priority sequence: finish in-flight then materiality. P2 accounts 73126, 73212, 73123, 73990', 'Open'),
]
B9 = Font(name='Segoe UI', size=9)
for i, (typ, acct, item, status) in enumerate(ITEMS, 1):
    r = 4 + i
    for j, v in enumerate((i, typ, acct, item, status), 1):
        c = oi.cell(r, j, v)
        c.font = B9
        c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=(j >= 4))
print(f'open items: {len(ITEMS)} rows')
wb.save('00_NA_Status_All_Accounts_11-Jun-2026.xlsx')
print('status workbook saved')
