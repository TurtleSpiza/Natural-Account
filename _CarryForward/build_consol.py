import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------- gather component lines ----------
def read_doc(path):
    wb=openpyxl.load_workbook(path, data_only=True); ws=wb['Sheet1']
    return [['PK',str(r[1]),str(r[2]),str(r[3]),str(r[4]),float(r[5]),str(r[6] or ''),str(r[7] or ''),str(r[8] or '')]
            for r in ws.iter_rows(values_only=True) if r and r[0]=='PK']

d45=read_doc('Document Line Table (45).xlsx')                 # PCard batch (Doc 1227681)
d46=[x for x in read_doc('Document Line Table (46).xlsx') if 'J12 recode' not in x[6]]  # 72111 audit minus J12

# WINC P1-P10 from audit workbook (cols start at B)
wb=openpyxl.load_workbook('73563_Coding_Review_LineItem_Audit.xlsx', data_only=True); ws=wb['Recode Journal']
winc=[]
for r in ws.iter_rows(values_only=True):
    vals=[('' if c is None else c) for c in r]
    s=[str(v) for v in vals]
    if 'PK' in s[:3]:
        i=s.index('PK'); seg=vals[i:i+9]
        if len(seg)>=6 and isinstance(seg[5],(int,float)):
            winc.append(['PK',str(seg[1]),str(seg[2]),str(seg[3]),str(seg[4]),float(seg[5]),str(seg[6]),str(seg[7]),str(seg[8])])

# WINC April aggregates (PK, dest -> amount)
apr={('PK000001','72111'):107.07,('PK000001','73512'):402.07,('PK000001','73564'):507.51,
     ('PK000086','72111'):50.06,('PK000086','73512'):19.70,('PK000086','73564'):101.20}
nar={'72111':('Recode WINC Apr non-stat fr 73563','Cleaning kitchen batteries','To 72111 Minor Equip & Supplies'),
     '73512':('Recode WINC Apr tearoom fr 73563','Coffee tea sugar lollies(Rsn B)','To 73512 Hospitality Non-FBT'),
     '73564':('Recode WINC Apr IT fr 73563','Headset chargers kbd adapter','To 73564 IT Equipment & Apps')}
winc_apr=[]
for (pk,dest),amt in apr.items():
    n1,n2,n3=nar[dest]
    winc_apr.append(['PK',pk,pk,'JOURNAL','73563',-round(amt,2),n1,n2,n3])
    winc_apr.append(['PK',pk,pk,'JOURNAL',dest,round(amt,2),n1,n2,n3])

# Reali net-movement (NA 72114)
reali=[
 ['PK','PK000012','PK000012','JOURNAL','72114',102.78,'Reali uniform recode fr 20151','PO702344 SI consol Mar-2026','72114 net-mvmt (confirm w Fin)'],
 ['PK','PK000433','PK000433','JOURNAL','72114',229.34,'Reali uniform recode fr 20151','PO702344 SI consol Mar-2026','72114 net-mvmt (confirm w Fin)'],
 ['PK','PK000435','PK000435','JOURNAL','72114',589.63,'Reali uniform recode fr 20151','PO702344 SI consol Mar-2026','72114 net-mvmt (confirm w Fin)'],
 ['PK','PK000003','PK000003','JOURNAL','72114',45.29,'Reali uniform recode fr 20151','PO702344 SI consol Mar-2026','72114 net-mvmt (confirm w Fin)'],
 ['PK','PK000402','PK000402','JOURNAL','72114',92.72,'Reali uniform recode fr 20151','PO702344 SI consol Mar-2026','72114 net-mvmt (confirm w Fin)'],
 ['PK','PK000001','PK000001','JOURNAL','72114',-1059.76,'Reali uniform recode fr 20151','PO702344 SI consol Mar-2026','72114 net-mvmt incl 1c round'],
]

streams=[('WINC 73563 P1-P10',winc),('WINC 73563 Apr-2026',winc_apr),
         ('72111 Minor Equip Audit',d46),('PCard Batch (Officeworks/Flair/Oasis)',d45),
         ('Reali Uniforms PK split',reali)]
for nm,ls in streams: print(nm,len(ls),'bal',round(sum(x[5] for x in ls),2))
allrows=[]
for nm,ls in streams:
    for x in ls: allrows.append([nm]+x)
print('TOTAL lines',len(allrows),'BATCH BALANCE',round(sum(x[6] for x in allrows),2))

# ---------- styling ----------
PANEL='1A2236'; NAVY='1F3864'; BLUE1='2F5597'; ACCENT='7C9CFF'; BAND='F4F6FA'; WHITE='FFFFFF'
GREEN_L='E2EFDA'; AMBER_L='FFF2CC'; RED_L='FFE0E0'; GREEN_D='375623'; AMBER='C8783C'; RED='C00000'; MUTE='595959'; GREY_H='D9D9D9'
def fill(c): return PatternFill('solid',fgColor=c)
def F(sz=10,bold=False,color='000000',name='Segoe UI'): return Font(name=name,size=sz,bold=bold,color=color)
def cen(w=False): return Alignment(horizontal='center',vertical='center',wrap_text=w)
def lft(w=True): return Alignment(horizontal='left',vertical='center',wrap_text=w)
def rgt(): return Alignment(horizontal='right',vertical='center')
thin=Side(style='thin',color='AAAAAA'); BORD=Border(top=thin,bottom=thin,left=thin,right=thin)
CUR='$#,##0.00;[Red]($#,##0.00);-'
out=Workbook()

# =================== JOURNAL SHEET ===================
js=out.active; js.title='Consolidated Journal'
js.sheet_view.showGridLines=False
hdr=['Stream','LDG','Account (PK)','Fund Account','Resource Group','Resource (NA)','Amount','Narrative 1','Narrative 2','Narrative 3']
title='CONSOLIDATED RECODE JOURNAL — TechOne GENJNL (PK Ledger)  |  Parks Branch 4090000  |  Prepared 3-Jun-2026'
js.merge_cells('A1:J1'); c=js['A1']; c.value=title; c.font=F(12,True,'FFFFFF'); c.fill=fill(PANEL); c.alignment=lft(False); js.row_dimensions[1].height=26
sub='40-char per narrative field. Dr positive, Cr negative. Batch must net to $0.00. Streams individually balanced.'
js.merge_cells('A2:J2'); c=js['A2']; c.value=sub; c.font=F(9,False,'D9D9D9'); c.fill=fill(NAVY); c.alignment=lft(False); js.row_dimensions[2].height=16
hr=4
for j,h in enumerate(hdr,1):
    cc=js.cell(hr,j,h); cc.font=F(10,True,'FFFFFF'); cc.fill=fill(NAVY); cc.alignment=cen(True); cc.border=BORD
r=hr+1; first_data=r
stream_color={'WINC 73563 P1-P10':'EAF1FB','WINC 73563 Apr-2026':'DCE8F8','72111 Minor Equip Audit':'F4F6FA','PCard Batch (Officeworks/Flair/Oasis)':'FBF0E4','Reali Uniforms PK split':'EDE7F6'}
for row in allrows:
    band=stream_color.get(row[0],'FFFFFF')
    for j,val in enumerate(row,1):
        cc=js.cell(r,j,val); cc.border=BORD; cc.font=F(9)
        cc.fill=fill(band)
        if j==7:
            cc.number_format=CUR; cc.alignment=rgt()
            if val<0: cc.font=F(9,color='C00000')
        elif j in (8,9,10): cc.alignment=lft(True); cc.font=F(8)
        elif j==1: cc.alignment=lft(True); cc.font=F(8,True,MUTE)
        else: cc.alignment=cen()
    r+=1
last_data=r-1
# balance row
js.cell(r,6,'BATCH NET').font=F(10,True,'FFFFFF'); js.cell(r,6).fill=fill(PANEL); js.cell(r,6).alignment=rgt()
bc=js.cell(r,7,f'=SUM(G{first_data}:G{last_data})'); bc.number_format=CUR; bc.font=F(11,True,'FFFFFF'); bc.fill=fill(PANEL); bc.alignment=rgt(); bc.border=BORD
for j in [1,2,3,4,5,8,9,10]: js.cell(r,j).fill=fill(PANEL)
js.row_dimensions[r].height=20
balance_row=r
widths=[26,5,13,13,13,12,13,30,30,30]
for j,w in enumerate(widths,1): js.column_dimensions[get_column_letter(j)].width=w
js.freeze_panes='A5'

# store ranges for cross-sheet formulas
G=f"'Consolidated Journal'!$G${first_data}:$G${last_data}"
STREAM=f"'Consolidated Journal'!$A${first_data}:$A${last_data}"
NA=f"'Consolidated Journal'!$F${first_data}:$F${last_data}"
PKR=f"'Consolidated Journal'!$C${first_data}:$C${last_data}"

# =================== SUMMARY SHEET ===================
sm=out.create_sheet('Summary',0); sm.sheet_view.showGridLines=False
sm.merge_cells('A1:H1'); c=sm['A1']; c.value='PARKS BRANCH 4090000 — CONSOLIDATED CODING-REVIEW JOURNAL  |  Reconciliation & Drill-Down'; c.font=F(13,True,'FFFFFF'); c.fill=fill(PANEL); c.alignment=lft(False); sm.row_dimensions[1].height=30
sm.merge_cells('A2:H2'); c=sm['A2']; c.value='Spero Karkalemis, Assistant Financial Analyst  •  Prepared 3-Jun-2026  •  All amounts GST-exclusive  •  TechOne PK ledger'; c.font=F(9,False,'D9D9D9'); c.fill=fill(NAVY); c.alignment=lft(False); sm.row_dimensions[2].height=16

# KPI cards
cards=[('Total lines',f'=COUNTA({STREAM})','#,##0'),
       ('Batch balance',f"='Consolidated Journal'!G{balance_row}",CUR),
       ('Review streams','5','#,##0'),
       ('Gross debits',f'=SUMIF({G},">0")',CUR),
       ('Gross credits',f'=SUMIF({G},"<0")',CUR),
       ('73563 moved out',f'=-SUMIF({NA},"73563",{G})',CUR),
       ('72111 moved out',f'=-SUMIF({NA},"72111",{G})',CUR),
       ('Outstanding flags','4','#,##0')]
r=4; col=1
for i,(lab,formula,fmt) in enumerate(cards):
    c0=col+(i%4)*2
    if i==4 and i>0: pass
    if i%4==0 and i>0: r+=3
    sm.merge_cells(start_row=r,start_column=c0,end_row=r,end_column=c0+1)
    sm.merge_cells(start_row=r+1,start_column=c0,end_row=r+1,end_column=c0+1)
    lc=sm.cell(r,c0,lab); lc.font=F(9,False,'FFFFFF'); lc.fill=fill(PANEL); lc.alignment=cen()
    vc=sm.cell(r+1,c0,formula); vc.font=F(16,True,ACCENT); vc.fill=fill(PANEL); vc.alignment=cen(); vc.number_format=fmt
    sm.cell(r,c0+1).fill=fill(PANEL); sm.cell(r+1,c0+1).fill=fill(PANEL)
    sm.row_dimensions[r].height=18; sm.row_dimensions[r+1].height=34
for j in range(1,9): sm.column_dimensions[get_column_letter(j)].width=15

# ---- Drill 1: by stream ----
r=12
def section(r,txt):
    sm.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8)
    c=sm.cell(r,1,txt); c.font=F(11,True,'FFFFFF'); c.fill=fill(NAVY); c.alignment=lft(False); sm.row_dimensions[r].height=20
    return r+1
r=section(r,'DRILL 1 — By review stream  (each stream must net to $0.00)')
for j,h in enumerate(['Review stream','Lines','Gross Dr','Gross Cr','Net (must=0)'],1):
    cc=sm.cell(r,j,h); cc.font=F(9,True,'FFFFFF'); cc.fill=fill(BLUE1); cc.alignment=cen(True); cc.border=BORD
r+=1
snames=['WINC 73563 P1-P10','WINC 73563 Apr-2026','72111 Minor Equip Audit','PCard Batch (Officeworks/Flair/Oasis)','Reali Uniforms PK split']
for i,s in enumerate(snames):
    band=BAND if i%2 else WHITE
    sm.cell(r,1,s).font=F(9); sm.cell(r,1).alignment=lft(True)
    sm.cell(r,2,f'=COUNTIF({STREAM},A{r})').number_format='#,##0'
    sm.cell(r,3,f'=SUMIFS({G},{STREAM},A{r},{G},">0")').number_format=CUR
    sm.cell(r,4,f'=SUMIFS({G},{STREAM},A{r},{G},"<0")').number_format=CUR
    sm.cell(r,5,f'=SUMIF({STREAM},A{r},{G})').number_format=CUR
    for j in range(1,6):
        cc=sm.cell(r,j); cc.border=BORD; cc.fill=fill(band)
        if j>=3: cc.alignment=rgt()
        if j in (2,): cc.alignment=cen()
        if j>1 and j!=5: cc.font=F(9)
    r+=1
tot=r
sm.cell(r,1,'TOTAL').font=F(9,True,'FFFFFF'); sm.cell(r,1).fill=fill(PANEL)
sm.cell(r,2,f'=SUM(B{tot-5}:B{tot-1})').number_format='#,##0'
sm.cell(r,3,f'=SUM(C{tot-5}:C{tot-1})').number_format=CUR
sm.cell(r,4,f'=SUM(D{tot-5}:D{tot-1})').number_format=CUR
sm.cell(r,5,f'=SUM(E{tot-5}:E{tot-1})').number_format=CUR
for j in range(1,6):
    cc=sm.cell(r,j); cc.fill=fill(PANEL); cc.font=F(9,True,'FFFFFF'); cc.border=BORD
    if j>=3: cc.alignment=rgt()
    if j==2: cc.alignment=cen()

# ---- Drill 2: by natural account (net movement) ----
r+=2
r=section(r,'DRILL 2 — By natural account  (net signed movement: negative = moved OUT, positive = moved IN)')
for j,h in enumerate(['NA','Account name','Net movement'],1):
    cc=sm.cell(r,j,h); cc.font=F(9,True,'FFFFFF'); cc.fill=fill(BLUE1); cc.alignment=cen(True); cc.border=BORD
r+=1
na_names={'73563':'Printing & Stationery','72111':'Minor Equipment & Supplies','72112':'Chemical, fertiliser & paints',
 '72113':'Safety Equipment & Medical','72114':'Uniforms','72222':'Signs','72231':'Landscape Supplies','72248':'Crusher Dust',
 '72249':'Topsoil','72311':'Other Plant & Equipment','72313':'Furniture & Fittings','72315':'Parks Assets',
 '73211':'Contractors','73220':'Traffic Control','73512':'Entertainment & Hospitality Non-FBT','73513':'Employee Reward & Recognition',
 '73564':'IT Equipment & Applications'}
nas=sorted(set(x[5] for x in allrows))
na_start=r
for i,na in enumerate(nas):
    band=BAND if i%2 else WHITE
    sm.cell(r,1,na).alignment=cen(); sm.cell(r,1).font=F(9)
    sm.cell(r,2,na_names.get(na,'')).alignment=lft(True); sm.cell(r,2).font=F(9)
    vc=sm.cell(r,3,f'=SUMIF({NA},A{r},{G})'); vc.number_format=CUR; vc.alignment=rgt()
    for j in range(1,4): sm.cell(r,j).border=BORD; sm.cell(r,j).fill=fill(band)
    r+=1
na_end=r-1
sm.cell(r,2,'TOTAL (must=0)').font=F(9,True,'FFFFFF'); sm.cell(r,2).fill=fill(PANEL); sm.cell(r,2).alignment=rgt()
sm.cell(r,3,f'=SUM(C{na_start}:C{na_end})').number_format=CUR; sm.cell(r,3).font=F(9,True,'FFFFFF'); sm.cell(r,3).fill=fill(PANEL); sm.cell(r,3).alignment=rgt()
sm.cell(r,1).fill=fill(PANEL)
for j in range(1,4): sm.cell(r,j).border=BORD

# ---- Drill 3: by PK (net movement across PKs) ----
r+=2
r=section(r,'DRILL 3 — By PK  (net signed movement: 0 = recode stayed within PK; non-zero = cross-PK relocation)')
for j,h in enumerate(['PK','Lines','Net movement'],1):
    cc=sm.cell(r,j,h); cc.font=F(9,True,'FFFFFF'); cc.fill=fill(BLUE1); cc.alignment=cen(True); cc.border=BORD
r+=1
pks=sorted(set(x[2] for x in allrows))
pk_start=r
for i,pk in enumerate(pks):
    band=BAND if i%2 else WHITE
    sm.cell(r,1,pk).alignment=cen(); sm.cell(r,1).font=F(9)
    sm.cell(r,2,f'=COUNTIF({PKR},A{r})').number_format='#,##0'; sm.cell(r,2).alignment=cen()
    vc=sm.cell(r,3,f'=SUMIF({PKR},A{r},{G})'); vc.number_format=CUR; vc.alignment=rgt()
    for j in range(1,4): sm.cell(r,j).border=BORD; sm.cell(r,j).fill=fill(band); 
    sm.cell(r,2).font=F(9)
    r+=1
pk_end=r-1
sm.cell(r,1,'TOTAL').font=F(9,True,'FFFFFF'); sm.cell(r,1).fill=fill(PANEL)
sm.cell(r,2,f'=SUM(B{pk_start}:B{pk_end})').number_format='#,##0'; sm.cell(r,2).alignment=cen()
sm.cell(r,3,f'=SUM(C{pk_start}:C{pk_end})').number_format=CUR; sm.cell(r,3).alignment=rgt()
for j in range(1,4): cc=sm.cell(r,j); cc.fill=fill(PANEL); cc.font=F(9,True,'FFFFFF'); cc.border=BORD

# =================== APRIL WINC DETAIL SHEET ===================
import csv
ad=out.create_sheet('April WINC Detail'); ad.sheet_view.showGridLines=False
ad.merge_cells('A1:I1'); c=ad['A1']; c.value='WINC INVOICE 9901782965 — APRIL 2026  |  Line-Item Classification  |  Parks accts 10179376 (PK000001) + 10428958 (PK000086)'; c.font=F(12,True,'FFFFFF'); c.fill=fill(PANEL); c.alignment=lft(False); ad.row_dimensions[1].height=26
ad.merge_cells('A2:I2'); c=ad['A2']; c.value='39 Parks lines, all currently coded 73563. Amounts GST-exclusive. GREEN = correctly in 73563; AMBER = miscoded, recode in consolidated journal.'; c.font=F(9,False,'D9D9D9'); c.fill=fill(NAVY); c.alignment=lft(False); ad.row_dimensions[2].height=16
ahdr=['PK','Item code','Description','Qty','Amount','Current','Suggested','Treatment','Status']
hr=4
for j,h in enumerate(ahdr,1):
    cc=ad.cell(hr,j,h); cc.font=F(10,True,'FFFFFF'); cc.fill=fill(NAVY); cc.alignment=cen(True); cc.border=BORD
r=hr+1; afirst=r
with open('april_classified.csv') as f:
    rows=list(csv.DictReader(f))
for row in rows:
    miss = row['status']=='MISCODE'
    band = AMBER_L if miss else GREEN_L
    vals=[row['pk'],row['item_code'],row['desc'],float(row['qty']),float(row['amount']),'73563',row['suggested_acct'],row['suggested_category'],row['status']]
    for j,val in enumerate(vals,1):
        cc=ad.cell(r,j,val); cc.border=BORD; cc.font=F(9); cc.fill=fill(band)
        if j==5: cc.number_format=CUR; cc.alignment=rgt()
        elif j==4: cc.alignment=cen()
        elif j in (1,6,7,9): cc.alignment=cen()
        else: cc.alignment=lft(True); cc.font=F(8)
        if j==9: cc.font=F(8,True, RED if miss else GREEN_D)
    r+=1
alast=r-1
# totals
ad.cell(r,4,'TOTAL').font=F(10,True,'FFFFFF'); ad.cell(r,4).fill=fill(PANEL); ad.cell(r,4).alignment=rgt()
tc=ad.cell(r,5,f'=SUM(E{afirst}:E{alast})'); tc.number_format=CUR; tc.font=F(10,True,'FFFFFF'); tc.fill=fill(PANEL); tc.alignment=rgt(); tc.border=BORD
for j in [1,2,3,6,7,8,9]: ad.cell(r,j).fill=fill(PANEL)
r+=1
ad.cell(r,4,'Miscoded (recoded out)').font=F(9,True,'FFFFFF'); ad.cell(r,4).fill=fill(AMBER); ad.cell(r,4).alignment=rgt()
mc=ad.cell(r,5,f'=SUMIF(I{afirst}:I{alast},"MISCODE",E{afirst}:E{alast})'); mc.number_format=CUR; mc.font=F(9,True,'FFFFFF'); mc.fill=fill(AMBER); mc.alignment=rgt(); mc.border=BORD
r+=1
ad.cell(r,4,'Correctly in 73563').font=F(9,True,'FFFFFF'); ad.cell(r,4).fill=fill(GREEN_D); ad.cell(r,4).alignment=rgt()
oc=ad.cell(r,5,f'=SUMIF(I{afirst}:I{alast},"OK",E{afirst}:E{alast})'); oc.number_format=CUR; oc.font=F(9,True,'FFFFFF'); oc.fill=fill(GREEN_D); oc.alignment=rgt(); oc.border=BORD
awidths=[10,11,42,6,12,9,11,34,10]
for j,w in enumerate(awidths,1): ad.column_dimensions[get_column_letter(j)].width=w
ad.freeze_panes='A5'

# =================== VERIFICATION & FINDINGS SHEET ===================
vf=out.create_sheet('Verification & Findings'); vf.sheet_view.showGridLines=False
vf.merge_cells('A1:F1'); c=vf['A1']; c.value='VERIFICATION & FINDINGS  |  Consolidated Coding-Review Journal  |  Parks Branch 4090000'; c.font=F(13,True,'FFFFFF'); c.fill=fill(PANEL); c.alignment=lft(False); vf.row_dimensions[1].height=28
vf.merge_cells('A2:F2'); c=vf['A2']; c.value='RED = blocker, resolve before upload  •  AMBER = confirm before upload  •  GREEN = verified clean  •  Prepared 3-Jun-2026'; c.font=F(9,False,'D9D9D9'); c.fill=fill(NAVY); c.alignment=lft(False); vf.row_dimensions[2].height=16

def vsection(r,txt):
    vf.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
    c=vf.cell(r,1,txt); c.font=F(11,True,'FFFFFF'); c.fill=fill(NAVY); c.alignment=lft(False); vf.row_dimensions[r].height=20
    return r+1
fhdr=['Flag','Stream / area','Item / ref','Finding','$ impact','Action required']
def fheader(r):
    for j,h in enumerate(fhdr,1):
        cc=vf.cell(r,j,h); cc.font=F(9,True,'FFFFFF'); cc.fill=fill(BLUE1); cc.alignment=cen(True); cc.border=BORD
    return r+1
def frow(r,flag,stream,ref,finding,impact,action):
    rag={'RED':(RED_L,RED),'AMBER':(AMBER_L,AMBER),'GREEN':(GREEN_L,GREEN_D)}[flag]
    cells=[flag,stream,ref,finding,impact,action]
    for j,val in enumerate(cells,1):
        cc=vf.cell(r,j,val); cc.border=BORD; cc.fill=fill(rag[0]); cc.font=F(9)
        if j==1: cc.alignment=cen(); cc.font=F(9,True,rag[1])
        elif j==5:
            cc.alignment=rgt()
            if isinstance(val,(int,float)): cc.number_format=CUR
            else: cc.alignment=cen()
        else: cc.alignment=lft(True)
        if j==4: cc.font=F(8)
        if j==6: cc.font=F(8)
    vf.row_dimensions[r].height=42
    return r+1

r=4
r=vsection(r,'1 — BLOCKER (resolve before upload)')
r=fheader(r)
r=frow(r,'RED','72111 Audit vs PCard Batch','Doc46 J12 / TE005091 vs Officeworks inv 626529478',
       'J12 recodes "20 desk risers" 72111 to 72313 ($363.64). The same 20 Otto risers are already split in the PCard Batch (Doc45) recode of Officeworks inv 626529478 ($398.18 = 20 Otto + 2 mesh + adapter). Booking both double-counts the riser relocation.',
       363.64,'CONFIRM TE005091 is the riser line of inv 626529478. If yes, J12 stays removed from the batch (already excluded here). If a separate purchase, re-add J12 (2 lines).')

r+=1
r=vsection(r,'2 — CONFIRM BEFORE UPLOAD')
r=fheader(r)
r=frow(r,'AMBER','PCard Batch (Flair Floral)','Flair Floral $77.27 / PK000001',
       'Reversal in Doc45 lifts from 72111, but the P-Card cover slip wrote the cost to 71111 (Full Time Staff). Reversal source must equal the live GL line.',
       77.27,'Confirm live GL coding: if it sits in 71111, change the reversal source from 72111 to 71111 before upload.')
r=frow(r,'AMBER','Reali Uniforms','PO702344 / inv PKLCC2026-0331',
       'Reali stream uses net-movement format (DR five PKs, CR PK000001); the four other streams use full-reversal. NA 72114 correct throughout, PK split is the only correction.',
       1059.76,'Confirm Finance accepts net-movement for this batch, or convert to full-reversal per their preference.')
r=frow(r,'AMBER','Travel — Gold Coast parking','TE004218 / 73541 to 73531',
       'City of Gold Coast parking $6.05, Craig Logan P-Card 16-Jun-2025, miscoded to 73541 Conferences. SEQ parking belongs in 73531 Local Travel. Not in this batch (need home PK).',
       6.05,'Provide Craig Logan home PK to build the 73541 to 73531 recode line.')
r=frow(r,'AMBER','Travel — Cabcharge','Jan-2026 statement $787.74',
       'Mixed local SEQ legs (73531) and interstate Melbourne/Geelong legs (73533). Needs per-traveller split. Not in this batch (need current GL coding + per-traveller PK + hotel-to-hotel destination).',
       787.74,'Provide current GL coding, per-traveller PK, and trip destinations to split 73531 vs 73533.')

r+=1
r=vsection(r,'3 — CARRY-FORWARD OPEN ITEM')
r=fheader(r)
r=frow(r,'AMBER','WINC 73563','inv 9901763737 / Feb-2026',
       'Hits GL in P9. PDF still absent from the pack, so no line-item classification possible. Flagged in prior findings, remains open.',
       1240.45,'Obtain the PDF for inv 9901763737 to complete the WINC 73563 audit trail.')

r+=1
r=vsection(r,'4 — VERIFIED CLEAN')
r=fheader(r)
r=frow(r,'GREEN','Batch integrity','All 5 streams',
       'Consolidated batch nets to $0.00. Each stream balances individually: WINC P1-P10, WINC Apr, 72111 Audit (minus J12), PCard Batch, Reali.','Nil','None.')
r=frow(r,'GREEN','Source docs balance','Doc45 / Doc46',
       'TechOne exports Doc File 1227681 (PCard) and 1226060 (72111 audit) both balance to $0.00 as extracted.','Nil','None.')
r=frow(r,'GREEN','Destination accounts','All NAs',
       'Every destination natural account exists in the Chart of Accounts and is UNRESTRICTED (72111/72112/72113/72114/72222/72231/72248/72249/72311/72313/72315/73211/73220/73512/73513/73563/73564).','Nil','None.')
r=frow(r,'GREEN','PK / service map','All PKs',
       'All PKs valid; narration service codes reconcile to the service-section map. Sub-PKs PK000445/PK000083/PK000493 not in the 101-row parent CSV but their services (20451/20392) are valid.','Nil','None.')
r=frow(r,'GREEN','April classification','inv 9901782965',
       '39 Parks lines reconciled: $1,676.26 total = $1,187.61 miscoded + $488.65 correctly in 73563. Aggregate recode matches line detail exactly.','Nil','None.')
r=frow(r,'GREEN','Travel taxonomy','73531-73541',
       'Confirmed: 73531 Local SEQ, 73532 Intrastate QLD, 73533 Interstate, 73534 Overseas, 73535 Tollway, 73541 Conferences (registration only). Gold Coast = SEQ.','Nil','None.')

vwidths=[8,24,26,62,12,46]
for j,w in enumerate(vwidths,1): vf.column_dimensions[get_column_letter(j)].width=w

out.save('/home/claude/out_consol.xlsx')
print('saved all sheets; na rows',len(nas),'pk rows',len(pks),'april rows',len(rows))
