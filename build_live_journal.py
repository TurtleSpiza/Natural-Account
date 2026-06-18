#!/usr/bin/env python3
"""Build 00_Live_Recode_Journal.xlsx, the live consolidated recode journal.

Parses every NA*_GENJNL_Recode.txt in the repo root, extracts the GENJNL PK/SL
ledger lines, and assembles the journal in canonical TechOne upload format
(matching the carried Consolidated Recode Journal): a Consolidated Journal sheet
(Stream | LDG | Account (PK) | Fund Account | Resource Group | Resource (NA) |
Amount | Narrative 1-3), a Summary, and a per-natural-account movement sheet.
Re-run whenever a recode changes - this is the live journal.

Repo tooling, excluded from the bundle manifest (like build_listing.py /
build_master.py). The xlsx it writes IS a manifested deliverable.
"""
from __future__ import annotations
import re, sys, os, csv
from collections import defaultdict
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

NAVY="1F3864"; BAND="EAEFF7"; WHITE="FFFFFF"
HF=PatternFill("solid", fgColor=NAVY); BANDF=PatternFill("solid", fgColor=BAND)
thin=Side(style="thin", color="BFBFBF"); BD=Border(thin,thin,thin,thin)
MONEY='#,##0.00;[Red]-#,##0.00'
NA_RE=re.compile(r'^[0-9][0-9A-Z]{4}$')
SET_RE=re.compile(r'^(SET\s+\d+|SUB-BATCH\s+[AB])\b[ :-]*(.*)', re.I)

LABELS={'NA72111':'72111 Minor Equipment','NA72312':'72312->72313 Furniture',
        'NA73513':'73513 Flair Floral','NA73533':'73533 Travel meal',
        'NA73563':'73563 WINC stationery','NA73564':'73564 IT Equipment',
        'NACarried3Jun':'Carried 3-Jun (Doc46+PCard)','NA72114':'72114 Reali uniforms',
        'NA73128':'73128->73512 TCB milk','NA73544':'73544 Training PK splits'}

# ---------- Evidence linkage ---------------------------------------------------
# Each line is traced to the actual filed support: an exact reference-token match
# against the account's evidence manifest, a paired-leg inheritance (a recode's
# reversal leg shares its partner's document), or the account's primary support
# file. The repo-relative path and SHA-256 come from 00_Bundle_Manifest.csv so a
# reviewer can open the exact file and integrity-check it.
EV_MANIFEST={'NA72111':'NA72111_Evidence_Manifest.csv','NA72114':'NA72114_Evidence_Manifest.csv',
        'NA72312':'NA72312_Evidence_Manifest.csv','NA73128':'NA73128_Evidence_Manifest.csv',
        'NA73433':'NA73433_Evidence_Manifest.csv','NA73513':'NA73513_Evidence_Manifest.csv',
        'NA73533':'NA73533_Evidence_Manifest.csv','NA73544':'NA73544_Evidence_Manifest.csv',
        'NA73563':'NA73563_Evidence_Manifest.csv','NA73564':'NA73564_Evidence_Manifest.csv',
        'NA7B213':'NA7B213_Evidence_Manifest.csv','NACarried3Jun':'NA72111_Evidence_Manifest.csv'}
# Per-account primary support for reference-less lines (internal recharges,
# consolidated standing-order reversals, bulk-audit lines). Real, manifested files.
EV_PRIMARY={'NA72111':'NA72111_Evidence/NA72111_Ledger_Full_9-Jun-2026.xlsx',
        'NA72114':'_Sources_9-Jun-2026/72114_GJ_Document_Reconstruction.xlsx',
        'NA72312':'NA72312_Evidence/NA72312_Ledger_Extract.xlsx',
        'NA73128':'NA73128_Evidence/NA73128_Ledger_Extract_10-Jun-2026.xlsx',
        'NA73433':'NA73433_Evidence_Manifest.csv',
        'NA73513':'NA73513_Evidence/NA73513_FlairFloral_coverslip_order.pdf',
        'NA73533':'NA73533_Evidence/NA73533_Travel_Support_Pack.pdf',
        'NA73544':'NA73544_Evidence_Manifest.csv',
        'NA73563':'NA73563_Evidence/NA73563_WINC_LineItem_Audit_P1-P10.xlsx',
        'NA73564':'NA73564_Evidence/NA73564_Ledger_Extract_9-Jun-2026.xlsx',
        'NA7B213':'NA7B213_Evidence/NA7B213_Ledger_4090000_18-Jun-2026.xlsx',
        'NACarried3Jun':'_CarryForward/72111_Doc46_Audit_source.xlsx'}
REF_RE=re.compile(r'(PIK\d{4,}|AUINV\w+|INV[-\s]?\d{3,}\w*|TE\d{4,}|SR\d{4,}|GJ\d{4,}|IJ\d{4,}|SC\d{3,}|PO\d{4,}|RL\d{4,}|\d{6,})')
EV_STOP=set('RECODE FROM MINOR EQUIP EQUIPMENT SUPPLIES NOTE SVCS SERVICE SERVICES PARK PARKS HOME '
            'TRNG TRAINING REVERSAL RECLASSIFI SYMPATHY MARKERS PAINT MEAL TRAVEL UNIFORM UNIFORMS '
            'UNIT UNITS COVER SLIP TAX FULL PARTIAL CONSOL BATCH JOURNAL LICENCE STORAGE'.split())
def _reftoks(*ss):
    out=set()
    for s in ss:
        s=re.sub(r'PK0*\d+','',str(s).upper())     # PK ids are not evidence references
        for m in REF_RE.findall(s):
            t=re.sub(r'[\s-]','',m); out.add(t)
            core=re.sub(r'\D','',t).lstrip('0')      # numeric core: INV-11404 == 00011404
            if len(core)>=4: out.add('N'+core)
    return out
def _vendtoks(*ss):
    out=set()
    for s in ss:
        for w in re.findall(r'[A-Z]{4,}', str(s).upper()):
            if w not in EV_STOP: out.add(w)
    return out
def load_ev_manifest(root: Path, path: str):
    """Reference tokens come ONLY from the file name and the dedicated reference/line
    columns - never from free-text note/status columns, whose prose can name a reference
    that is in fact still a GAP. Notes contribute vendor words only (a +1 tiebreak)."""
    rows=[]; fp=root/path
    if not fp.exists(): return rows
    REF_COLS={'reference','line','line_ref','ledger_ref','invoiceno','invoice','ref'}
    with fp.open(newline='', encoding='utf-8', errors='replace') as fh:
        rd=csv.reader(fh)
        try: hdr=next(rd)
        except StopIteration: return rows
        low=[h.lower().strip() for h in hdr]
        fi=low.index('file') if 'file' in low else (low.index('evidence_file') if 'evidence_file' in low else 0)
        ref_idx=[i for i,h in enumerate(low) if h in REF_COLS]
        for r in rd:
            if not any(r): continue
            f=r[fi] if fi<len(r) else ''
            if not f or f.strip().upper()=='GAP': continue   # skip GAP/placeholder rows
            base=os.path.basename(f)
            refcells=[r[i] for i in ref_idx if i<len(r)]
            rows.append({'file':f,'rt':_reftoks(base,*refcells),'vt':_vendtoks(base,*r)})
    return rows
def build_file_index(root: Path):
    """basename -> [repo-relative paths] and path -> sha256, from the bundle manifest."""
    base2paths=defaultdict(list); sha={}
    mp=root/'00_Bundle_Manifest.csv'
    if mp.exists():
        with mp.open(newline='', encoding='utf-8') as fh:
            rd=csv.reader(fh); next(rd, None)
            for row in rd:
                if len(row)<2: continue
                sha[row[0]]=row[1]; base2paths[os.path.basename(row[0])].append(row[0])
    return base2paths, sha
def resolve_path(file_str: str, acct: str, base2paths):
    """Map a manifest 'file' value (often a basename) to its real repo-relative path."""
    bn=os.path.basename((file_str or '').strip())
    paths=base2paths.get(bn, [])
    if not paths: return file_str
    if len(paths)==1: return paths[0]
    for p in paths:                       # disambiguate to the account's own copy
        if acct in p or acct[2:] in p: return p
    return paths[0]
def resolve_evidence(acct, n1, n2, n3, ev_cache, base2paths, sha):
    rows=ev_cache.get(acct, [])
    jr=_reftoks(n1,n2,n3); jv=_vendtoks(n1,n2,n3)
    best=None; bs=0
    for row in rows:
        score=2*len(jr & row['rt']) + len(jv & row['vt'])
        if score>bs: bs=score; best=row
    if best and (jr & best['rt']):        # assert a file only on a shared reference token
        rp=resolve_path(best['file'], acct, base2paths)
        if sha.get(rp):                   # and only if it is a real manifested deliverable
            return (rp, sha[rp], 'exact-ref')
    prim=EV_PRIMARY.get(acct, EV_MANIFEST.get(acct,''))
    return (prim, sha.get(prim,''), 'account-source')

def parse(path: Path, label: str):
    stream=label
    for raw in path.read_text(encoding="utf-8", errors="replace").splitlines():
        m=SET_RE.match(raw.strip())
        if m:
            tail=re.split(r'[(.]', m.group(2))[0].strip().strip('-').strip()
            stream=f"{label} - {m.group(1).upper()}" + (f" {tail[:28]}" if tail else "")
        if "\t" in raw and "|" not in raw:
            # Tab-delimited GENJNL with a leading Line No column (NA73128 onward).
            p=[x.strip() for x in raw.split("\t")]
            if p and p[0].isdigit():
                p=p[1:]
        elif "|" in raw:
            p=[x.strip() for x in raw.split("|")]
        else:
            continue
        if len(p)<6 or p[0] not in ("PK","SL"): continue
        if not NA_RE.match(p[4]): continue
        try: amt=float(p[5])
        except ValueError: continue
        n1,n2,n3=(p[6:9]+["","",""])[:3]
        yield (stream,p[0],p[1],p[2],p[3],p[4],amt,n1,n2,n3)

def style_header(ws, row, ncols, size=10):
    for c in range(1,ncols+1):
        x=ws.cell(row,c); x.font=Font(name="Segoe UI",bold=True,color=WHITE,size=size)
        x.fill=HF; x.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); x.border=BD

def main(argv):
    root=Path(argv[1]).resolve() if len(argv)>1 else Path(__file__).resolve().parent
    files=sorted(root.glob("NA*_GENJNL_Recode.txt"))
    base2paths, sha = build_file_index(root)
    ev_cache={a:load_ev_manifest(root,p) for a,p in {**EV_MANIFEST}.items()}
    lines=[]; ev=[]; per_file={}; per_na={}
    for f in files:
        acct=f.name.split("_")[0]
        lab=LABELS.get(acct, acct)
        recs=list(parse(f, lab))
        per_file[acct]={"lines":len(recs),"net":round(sum(r[6] for r in recs),2)}
        for r in recs:
            lines.append(r); per_na[r[5]]=per_na.get(r[5],0.0)+r[6]
            ev.append(list(resolve_evidence(acct, r[7], r[8], r[9], ev_cache, base2paths, sha)))
    # Pair-propagation: a recode's reversal leg inherits its partner's exact document,
    # but only across a BALANCED ADJACENT PAIR (neighbouring rows, same stream, equal
    # |amount|, opposite sign) - the off/to shape. Adjacency avoids cross-item bleed in
    # the long mixed streams, where unrelated items can share an amount.
    is_exact=[e[2]=='exact-ref' for e in ev]
    for i in range(len(lines)):
        if is_exact[i]: continue
        for j in (i-1, i+1):
            if not (0<=j<len(lines) and is_exact[j]): continue
            a,b=lines[i],lines[j]
            if a[0]==b[0] and round(abs(a[6]),2)==round(abs(b[6]),2) and (a[6]>0)!=(b[6]>0):
                ev[i]=[ev[j][0], ev[j][1], 'paired-leg']; break
    gross_dr=round(sum(r[6] for r in lines if r[6]>0),2)
    gross_cr=round(sum(r[6] for r in lines if r[6]<0),2)
    net=round(sum(r[6] for r in lines),2)
    streams=sorted({r[0] for r in lines})

    wb=openpyxl.Workbook()
    # ---------- Summary ----------
    ss=wb.active; ss.title="Summary"
    ss["A1"]="PARKS BRANCH 4090000 - LIVE CONSOLIDATED RECODE JOURNAL"; ss["A1"].font=Font(name="Segoe UI",bold=True,size=13,color=NAVY)
    ss["A2"]=("Batch CN-23167. TechOne GENJNL (PK ledger). Dr positive, Cr negative. Regenerated by build_live_journal.py. "
              "Every line is traced to its filed support (Evidence File + SHA-256) on the Consolidated Journal sheet.")
    ss["A2"].font=Font(name="Segoe UI",size=9,italic=True)
    ss.append([]);
    ss.append(["Total lines","Batch balance","Streams","Account recodes","Gross debits","Gross credits"])
    style_header(ss,4,6)
    ss.append([len(lines),net,len(streams),len(per_file),gross_dr,gross_cr])
    for c in range(1,7):
        x=ss.cell(5,c); x.font=Font(name="Segoe UI",size=11,bold=(c in(1,2))); x.border=BD; x.alignment=Alignment(horizontal="center")
    for c in (2,5,6): ss.cell(5,c).number_format=MONEY
    ss.cell(5,2).font=Font(name="Segoe UI",size=11,bold=True,color=("C00000" if abs(net)>0.005 else "1F7A1F"))
    # per-account block
    ss.append([]); r=7
    ss.cell(r,1,"Per account recode"); ss.cell(r,1).font=Font(name="Segoe UI",bold=True,size=11,color=NAVY); r+=1
    ss.append(["Account","Lines","Net (=0.00)"]); style_header(ss,r,3); r+=1
    for a in sorted(per_file):
        d=per_file[a]; ss.cell(r,1,a); ss.cell(r,2,d["lines"]); nc=ss.cell(r,3,d["net"]); nc.number_format=MONEY
        for c in range(1,4): ss.cell(r,c).font=Font(name="Segoe UI",size=9); ss.cell(r,c).border=BD
        if r%2==0:
            for c in range(1,4): ss.cell(r,c).fill=BANDF
        r+=1
    for c,w in zip("ABCDEF",[16,15,10,16,14,14]): ss.column_dimensions[c].width=w

    # ---------- Consolidated Journal (canonical upload format) ----------
    js=wb.create_sheet("Consolidated Journal")
    js["A1"]="CONSOLIDATED RECODE JOURNAL - CN-23167 (upload format)"; js["A1"].font=Font(name="Segoe UI",bold=True,size=12,color=NAVY)
    js["A2"]=("40-char narrative cap; PK preserved unless stated; batch nets $0.00. Store Account/Resource as text. "
              "Evidence File = repo-relative path to the actual filed support; SHA-256 from 00_Bundle_Manifest.csv; "
              "Basis: exact-ref (document number matched) / paired-leg (reversal of a matched line) / account-source (ledger or audit source for reference-less lines).")
    js["A2"].font=Font(name="Segoe UI",size=8,italic=True)
    cols=["Stream","LDG","Account (PK)","Fund Account","Resource Group","Resource (NA)","Amount",
          "Narrative 1","Narrative 2","Narrative 3","Evidence File","Evidence SHA-256","Evidence Basis"]
    hrow=4
    for c,t in enumerate(cols,1): js.cell(hrow,c,t)
    style_header(js,hrow,len(cols),size=9)
    for i,(r,e) in enumerate(zip(lines,ev), start=hrow+1):
        row=list(r)+[e[0], e[1], e[2]]
        for c,v in enumerate(row,1):
            x=js.cell(i,c,v); x.font=Font(name="Segoe UI",size=8); x.border=BD
            if c==7: x.number_format=MONEY; x.alignment=Alignment(horizontal="right")
            if c in (3,4,6,11,12): x.number_format='@'
            if i%2==0: x.fill=BANDF
    js.freeze_panes=f"H{hrow+1}"
    nr=hrow+1+len(lines)
    js.cell(nr,6,"BATCH NET").font=Font(name="Segoe UI",bold=True)
    nc=js.cell(nr,7,f"=SUM(G{hrow+1}:G{hrow+len(lines)})"); nc.font=Font(name="Segoe UI",bold=True); nc.number_format=MONEY
    for c,w in zip("ABCDEFGHIJKLM",[34,6,13,13,15,14,12,30,30,24,46,30,13]): js.column_dimensions[c].width=w

    # ---------- Net movement by NA ----------
    ms=wb.create_sheet("NA Movement")
    ms["A1"]="Net movement by natural account (+in / -out)"; ms["A1"].font=Font(name="Segoe UI",bold=True,size=12,color=NAVY)
    ms.append([]); ms.append(["Natural account","Net movement"]); style_header(ms,3,2)
    r=4
    for na in sorted(per_na):
        ms.cell(r,1,na).number_format='@'; nc=ms.cell(r,2,round(per_na[na],2)); nc.number_format=MONEY
        for c in (1,2): ms.cell(r,c).font=Font(name="Segoe UI",size=9); ms.cell(r,c).border=BD
        r+=1
    ms.cell(r,1,"GRAND NET").font=Font(name="Segoe UI",bold=True); gc=ms.cell(r,2,round(sum(per_na.values()),2)); gc.font=Font(name="Segoe UI",bold=True); gc.number_format=MONEY
    for c,w in zip("AB",[18,16]): ms.column_dimensions[c].width=w

    out=root/"00_Live_Recode_Journal.xlsx"; wb.save(out)
    print(f"Wrote {out.name}: {len(lines)} lines, {len(per_file)} account recodes, {len(streams)} streams.")
    print(f"Gross Dr {gross_dr:,.2f} | Gross Cr {gross_cr:,.2f} | NET {net:+.2f}")
    bad=[(a,d['net']) for a,d in per_file.items() if abs(d['net'])>0.005]
    print("Per-account net:", "ALL ZERO" if not bad else f"NON-ZERO {bad}")
    from collections import Counter as _C
    bc=_C(e[2] for e in ev)
    print("Evidence linkage:", dict(bc), f"| lines with a file ref: {sum(1 for e in ev if e[0])}/{len(ev)}")
    return 0

if __name__=="__main__":
    raise SystemExit(main(sys.argv))
