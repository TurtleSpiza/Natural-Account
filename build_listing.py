import pandas as pd, json, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
sm=json.load(open('.claude/skills/lcc-coding-review/svc_map.json'))['service_codes']
def svc_info(s): r=sm.get(s); return (r.get('pk',''),r.get('section','')) if r else ('','')
def vendor(d): return re.split(r'-General Expenses|-Purchase Card',str(d))[0].strip()[:34]
def load_acct(path,na):
    df=pd.read_excel(path,engine='calamine',header=5,dtype=object)
    df=df[df['Account'].astype(str).str.match(rf'^\d-\d{{5}}-{na}$').fillna(False)].copy()
    df['amt']=pd.to_numeric(df['Transaction Amount'],errors='coerce'); return df
O72111={
 ('TE005091',473.64):('G','G','R','G','G','R','Kirsty Quinn','Officeworks 626529478. Recode OUT per Doc45 (risers 398.18>72313, diary 43.64>73563, adapter 31.82>73564). Doc46 J12 duplicate DROPPED; RED gate resolved.','72313/73563/73564'),
 ('TE005080',11.36):('G','G','A','G','R','R','Kirsty Quinn','Coles milk $12.50 GST-FREE booked $11.36 = $1.14 phantom GST credit. Gross GL to $12.50, reverse ITC. Amenity.','gross-up $12.50'),
 ('TE005080',73.35):('G','G','G','G','G','G','Kirsty Quinn','Choice storage tubs (ABN 45 647 301 764) comms-room WHS storage. Taxable, GST $7.34, booked $73.35 correct. Clears all four limbs.',''),
 ('TE005250',6.10):('G','G','A','G','G','A','Craig Logan','Woolworths milk GST-free booked full = correct. Amenity.',''),
 ('TE005531',13.80):('G','G','A','G','G','A','Craig Logan','Coles milk GST-free booked full = correct. Amenity.',''),
 ('TE004477',71.00):('R','P','A','G','G','R','card ...0429','Bunnings (ABN 26 008 672 179) GST correct. Park materials > Parks Services svc 20392/PK000022.','PK000022'),
 ('TE004477',85.12):('R','P','A','A','R','R','card ...0429','Steel Post & Rail PAYMENT RECEIPT not tax invoice; $2.02 finance charge input-taxed, GST over-claimed. > PK000022.','PK000022'),
 ('TE004477',681.82):('R','P','A','A','P','R','card ...0429','Play Safety Australia, receipt not sighted. Park materials > PK000022.','PK000022'),
 ('TE004476',259.89):('A','P','A','G','G','A','card ...0429','Bunnings tool trolley+bolts (ABN 26 008 672 179) GST $26.00 correct. Confirm Depots vs Parks Services (batch also on 20392); P&A check on $226 trolley.',''),
 ('TE004476',40.38):('A','P','A','G','G','A','card ...0429','Bunnings cleaning/cable-ties GST $4.04 correct. Confirm Depots vs Parks Services; cleaning 72111 vs 73121.',''),
 ('TE004476',32.66):('G','P','P','A','P','A','card ...0429','On Parks Services svc 20392/PK000022 (correct split). Receipt unsighted.',''),
 ('TE004476',126.46):('G','P','P','A','P','A','card ...0429','On Parks Services svc 20392/PK000022 (correct split). Receipt unsighted.',''),
}
FOOD=re.compile(r'COLES|WOOLWORTHS|IGA|ALDI|BAKERS|7.?ELEVEN',re.I)
SEED={'72312':('G','G','R','G','G','R','Recode to 72313 Furniture & Fittings; held on Finance format. See Verification Record.'),
 '73533':('G','G','G','G','G','G','Interstate travel, correct in 73533. See Verification Record.'),
 '73563':('G','P','A','A','A','A','WINC catch-all; $6,354.47 miscoded out, 9 WINC invoices unsighted, refreshment lines incl GST-free food. See Verification Record.'),
 '73564':('A','P','A','A','A','A','IT Equipment; see Verification Record (PK attribution, 73566 software recode, evidence/tax gaps).')}
OV={'73533':{('TE005080',288.18):('G','G','G','G','A','A','Booking.com ibis offshore: no ABN, not a tax invoice; $28.82 ITC unsubstantiated pending Accor tax invoice.',''),
   ('TE005080',6.55):('G','G','A','A','A','A','7-Eleven sundry, receipt unsighted; likely meal>73511/73512; GST-free food possible.','')},
 '73563':{('GJ075544',-448.22):('G','G','G','G','G','G','EOY 2024-25 accrual reversal, normal, untouched.','')},
 '73564':{('TE005499',36.14):('A','P','R','A','A','R','Anthropic Claude (US) software>73566; offshore digital, verify no AU GST.','73566'),
   ('INAU004378',132.00):('A','P','R','A','A','R','License bundle>73566; ref unsighted.','73566'),
   ('INAU004378',342.00):('A','P','R','A','A','R','Pro license bundle>73566; ref unsighted.','73566'),
   ('GJ076194',372.00):('R','P','R','A','A','R','Body2 sub>73566; off-home PK; unsighted.','73566'),
   ('GJ075813',372.00):('R','P','R','A','A','R','Body2 sub>73566; off-home PK; unsighted.','73566'),
   ('INV000555',568.31):('A','P','R','G','G','R','Esri ArcGIS Creator>73566; invoice sighted compliant.','73566'),
   ('INV-13660',None):('A','P','G','G','G','A','Apple iPad fleet sighted compliant; P&A register (advisory).',''),
   ('INV-17543',None):('A','P','G','G','G','A','Apple iPad fleet sighted compliant; P&A register (advisory).','')}}
def build(df,na):
    rows=[]; seed=SEED.get(na); ov=OV.get(na,{})
    for _,r in df.iterrows():
        acct=str(r['Account']); svc=acct.split('-')[1]; pk,section=svc_info(svc)
        det=r['Details']; ven=vendor(det); ref=str(r['Reference']); amt=r['amt']; has=str(r['Has Attachment']).upper()=='Y'
        card=''; rc=''
        if na=='72111':
            o=O72111.get((ref,round(amt,2)) if amt==amt else (ref,None))
            if o: l1,l1e,l2,l3,l4,ovr,card,note,rc=o
            else:
                l1=l1e=l2='P'; l3='A' if has else 'R'
                if FOOD.search(str(det)): l4='A'; note='GST-free food? verify booked ex vs receipt.'
                elif ven.upper().startswith('SQ *'): l4='A'; note='Square trader: verify ABN/GST-registration.'
                else: l4='P'; note=''
                ovr='P' if l4=='P' else 'A'
        else:
            hit=None
            for (kref,kamt),val in ov.items():
                if kref in ref and (kamt is None or (amt==amt and abs(amt-kamt)<0.01)): hit=val; break
            if hit: l1,l1e,l2,l3,l4,ovr,note,rc=hit
            else: l1,l1e,l2,l3,l4,ovr,note=seed
        rows.append(dict(Account=na,Svc=svc,PK=pk,Section=section,Period=r['Period'],Date=str(r['Date'])[:10],
            Reference=ref,Vendor=ven,Cardholder=card,Details=str(det)[:58],
            ExGST=round(amt,2) if amt==amt else '',InclX11=round(amt*1.1,2) if amt==amt else '',Attach='Y' if has else 'N',
            L1_SvcPK=l1,L1_Emp=l1e,L2_NA=l2,L3_Evid=l3,L4_Tax=l4,Overall=ovr,Finding=note,RecodeTo=rc))
    return pd.DataFrame(rows)
accts=[('72111','Minor Equipment & Supplies','NA72111_Evidence/NA72111_Ledger_Full_9-Jun-2026.xlsx'),
 ('72312','Office & Library Equipment','NA72312_Evidence/NA72312_Ledger_Extract.xlsx'),
 ('73533','Interstate Travel & Accommodation','NA73533_Evidence/NA73533_Ledger_Extract.xlsx'),
 ('73563','Printing & Stationery','NA73563_Evidence/NA73563_Ledger_Extract.xlsx'),
 ('73564','IT Equipment & Applications','NA73564_Evidence/NA73564_Ledger_Extract_9-Jun-2026.xlsx')]
data={na:(nm,build(load_acct(p,na),na)) for na,nm,p in accts}
# render
NAVY="1F4E79"; HDR=Font(name="Segoe UI",bold=True,color="FFFFFF",size=10); B=Font(name="Segoe UI",size=10)
band=PatternFill("solid",fgColor="F2F2F2"); RAG={'G':("C6EFCE","006100"),'A':("FFEB9C","9C6500"),'R':("FFC7CE","9C0006"),'P':("D9D9D9","595959")}
thin=Side(style="thin",color="D9D9D9"); bd=Border(thin,thin,thin,thin); ctr=Alignment(horizontal="center",vertical="center"); lft=Alignment(horizontal="left",vertical="center")
def cnt(s,*c): return int(s.isin(c).sum())
wb=Workbook(); sw=wb.active; sw.title="Summary"
sw["A1"]="Parks 4090000 NA Review — Running Transaction Listing"; sw["A1"].font=Font(name="Segoe UI",bold=True,size=13,color=NAVY)
for j,h in enumerate(["Account","Lines","$ ex-GST","Reviewed %","L1 R/A","L2 R/A","L3 R/A","L4 R/A","Overall R/A/P"],1):
    c=sw.cell(3,j,h); c.font=HDR; c.fill=PatternFill("solid",fgColor=NAVY); c.alignment=ctr; c.border=bd
rr=4; gt=0; gl=0
for na,nm,_ in accts:
    df=data[na][1]; tot=pd.to_numeric(df['ExGST'],errors='coerce').sum(); gt+=tot; gl+=len(df); rev=int((df['Overall']!='P').sum())
    vals=[f"{na} {nm}",len(df),round(tot,2),f"{rev/len(df)*100:.0f}%",
          f"{cnt(df['L1_SvcPK'],'R')} / {cnt(df['L1_SvcPK'],'A')}",f"{cnt(df['L2_NA'],'R')} / {cnt(df['L2_NA'],'A')}",
          f"{cnt(df['L3_Evid'],'R')} / {cnt(df['L3_Evid'],'A')}",f"{cnt(df['L4_Tax'],'R')} / {cnt(df['L4_Tax'],'A')}",
          f"{cnt(df['Overall'],'R')} / {cnt(df['Overall'],'A')} / {cnt(df['Overall'],'P')}"]
    for j,v in enumerate(vals,1):
        c=sw.cell(rr,j,v); c.font=B; c.alignment=lft if j==1 else ctr; c.border=bd
        if j==3: c.number_format='#,##0.00'
    rr+=1
for na,nm,note in [('73140','Pre-Employment Background Checks','no line ledger; SE2 $3,703.03; see Verification Record'),('72114','Uniforms','partial; GJ reconstruction in _Sources; see Scope Note')]:
    sw.cell(rr,1,f"{na} {nm}").font=B; sw.cell(rr,1).border=bd; sw.cell(rr,9,note).font=Font(name="Segoe UI",italic=True,size=9,color="595959")
    for j in range(2,10): sw.cell(rr,j).border=bd
    rr+=1
sw.cell(rr,1,"TOTAL line-level").font=Font(name="Segoe UI",bold=True); sw.cell(rr,2,gl).font=Font(name="Segoe UI",bold=True)
sw.cell(rr,3,round(gt,2)).font=Font(name="Segoe UI",bold=True); sw.cell(rr,3).number_format='#,##0.00'
for j in range(1,10): sw.cell(rr,j).border=bd
for j,w in enumerate([40,8,14,12,11,11,11,11,16],1): sw.column_dimensions[get_column_letter(j)].width=w
lg=wb.create_sheet("Legend")
for i,(a,b) in enumerate([("Column/code","Meaning"),("L1 Svc/PK","service maps to correct PK and section"),("L1 Emp","cardholder belongs to that service/PK"),("L2 NA","natural account correct vs Chart"),("L3 Evid","own document sighted, sufficient"),("L4 Tax","ABN+GST math+booked GST status reconciles to document"),("Overall","worst of five; GREEN only when all pass"),("G","pass"),("A","confirm"),("R","blocker/error"),("P","pending")],1):
    ca=lg.cell(i,1,a); cb=lg.cell(i,2,b); ca.font=Font(name="Segoe UI",bold=(i==1),size=10); cb.font=Font(name="Segoe UI",bold=(i==1),size=10)
    if a in RAG: f,fc=RAG[a]; ca.fill=PatternFill("solid",fgColor=f); ca.font=Font(name="Segoe UI",bold=True,color=fc); ca.alignment=ctr
lg.column_dimensions["A"].width=14; lg.column_dimensions["B"].width=90
cols=[("Account",8),("Svc",6),("PK",10),("Section",22),("Period",6),("Date",11),("Reference",13),("Vendor",24),("Cardholder",15),("Details",44),("ExGST",10),("InclX11",10),("Attach",6),("L1 Svc/PK",9),("L1 Emp",8),("L2 NA",7),("L3 Evid",8),("L4 Tax",7),("Overall",8),("Finding / note",66),("Recode To",14)]
ragc={"L1 Svc/PK":"L1_SvcPK","L1 Emp":"L1_Emp","L2 NA":"L2_NA","L3 Evid":"L3_Evid","L4 Tax":"L4_Tax","Overall":"Overall"}
fm={"Account":"Account","Svc":"Svc","PK":"PK","Section":"Section","Period":"Period","Date":"Date","Reference":"Reference","Vendor":"Vendor","Cardholder":"Cardholder","Details":"Details","ExGST":"ExGST","InclX11":"InclX11","Attach":"Attach","Finding / note":"Finding","Recode To":"RecodeTo"}
for na,nm,_ in accts:
    df=data[na][1]; tot=pd.to_numeric(df['ExGST'],errors='coerce').sum(); ws=wb.create_sheet(na)
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(cols))
    t=ws.cell(1,1,f"NA {na} {nm} — line-level review ({len(df)} lines, ${tot:,.2f} ex-GST)"); t.font=Font(name="Segoe UI",bold=True,size=12,color="FFFFFF"); t.fill=PatternFill("solid",fgColor=NAVY); ws.row_dimensions[1].height=22
    for j,(n,w) in enumerate(cols,1):
        c=ws.cell(2,j,n); c.font=HDR; c.fill=PatternFill("solid",fgColor=NAVY); c.alignment=ctr; c.border=bd; ws.column_dimensions[get_column_letter(j)].width=w
    ws.row_dimensions[2].height=28; r=3
    for _,row in df.iterrows():
        for j,(n,w) in enumerate(cols,1):
            if n in ragc:
                code=row[ragc[n]]; c=ws.cell(r,j,code); f,fc=RAG.get(code,("FFFFFF","000000")); c.fill=PatternFill("solid",fgColor=f); c.font=Font(name="Segoe UI",bold=True,color=fc,size=10); c.alignment=ctr
            else:
                val=row[fm[n]]; c=ws.cell(r,j,val); c.font=B; c.alignment=ctr if n in("Period","Date","Attach","ExGST","InclX11","Svc") else lft
                if n in("ExGST","InclX11") and val!='': c.number_format='#,##0.00'
                if r%2==0: c.fill=band
            c.border=bd
        r+=1
    ws.freeze_panes="G3"; ws.auto_filter.ref=f"A2:{get_column_letter(len(cols))}{r-1}"
wb.save("00_Running_Transaction_Listing.xlsx")
rev72=int((data['72111'][1]['Overall']!='P').sum())
print(f"listing rebuilt. 72111 reviewed lines: {rev72}. total {gl} lines ${gt:,.2f}")
