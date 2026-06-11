#!/usr/bin/env python3
"""Build 00_NAReview_Master.xlsx: one workbook, three tiers, live rollups.

Self-contained and repo-runnable. Reads:
  00_Running_Transaction_Listing.xlsx   (line tier; the per-line verdicts)
  00_Account_Review_Register.xlsx        (account tier source rows)
  00_Parks_4090000_NAReview_Tracker.xlsx (programme tier source rows)

Emits 00_NAReview_Master.xlsx with tabs Tracker / Register / Summary / Legend /
one line tab per account, and LIVE intra-workbook formulas:
  line criteria  ->  Register $ / RAG / Reviewed$  ->  Tracker Reviewed$ / Untested$

NOTE: the Register RAG, the Register/Tracker totals and the Tracker dashboard are
formulas. Open in Excel or recalc with LibreOffice
(soffice --headless --convert-to xlsx) then read back data_only to verify before
retiring the three source files. Realigned 72111 reviewed = sum of non-PENDING
lines (Overall<>"P").
"""
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

LISTING = "00_Running_Transaction_Listing.xlsx"
REGISTER = "00_Account_Review_Register.xlsx"
TRACKER = "00_Parks_4090000_NAReview_Tracker.xlsx"
LINE_NAS = ["72111", "72312", "73533", "73563", "73564", "72114", "73511", "73128"]

NAVY = "1F4E79"
HDR = Font(name="Segoe UI", bold=True, color="FFFFFF", size=10)
BASE = Font(name="Segoe UI", size=10)
BOLD = Font(name="Segoe UI", bold=True, size=10)
band = PatternFill("solid", fgColor="F2F2F2")
RAG = {"G": ("C6EFCE", "006100"), "A": ("FFEB9C", "9C6500"),
       "R": ("FFC7CE", "9C0006"), "P": ("D9D9D9", "595959"),
       "J": ("A9D08E", "375623")}
RAGWORD = {"GREEN": "G", "AMBER": "A", "RED": "R", "PENDING": "P", "CLEARED": "J"}
thin = Side(style="thin", color="D9D9D9")
bd = Border(thin, thin, thin, thin)
ctr = Alignment(horizontal="center", vertical="center")
lft = Alignment(horizontal="left", vertical="center")

LCOLS = ["Account", "Svc", "PK", "Section", "Period", "Date", "Reference", "Vendor",
         "Cardholder", "Details", "ExGST", "InclX11", "Attach", "L1 Svc/PK", "L1 Emp",
         "L2 NA", "L3 Evid", "L4 Tax", "Overall", "Finding / note", "Recode To", "Journal"]
LWID = [8, 6, 10, 22, 6, 11, 13, 24, 15, 44, 10, 10, 6, 9, 8, 7, 8, 7, 8, 66, 14, 30]
RAGCOLS = {"L1 Svc/PK", "L1 Emp", "L2 NA", "L3 Evid", "L4 Tax", "Overall"}
# in the master line tabs: K=ExGST(11) N=L1(14) O=L1e(15) P=L2(16) Q=L3(17) R=L4(18) S=Overall(19)


def read_listing():
    wb = load_workbook(LISTING, data_only=True)
    out = {}
    for na in LINE_NAS:
        ws = wb[na]
        hdr = {ws.cell(2, c).value: c for c in range(1, ws.max_column + 1)}
        rows = []
        for r in range(3, ws.max_row + 1):
            if ws.cell(r, hdr["Reference"]).value is None and ws.cell(r, hdr["ExGST"]).value is None:
                continue
            rows.append([ws.cell(r, hdr[c]).value for c in LCOLS])
        out[na] = pd.DataFrame(rows, columns=LCOLS)
    return out


def rag_cf(ws, rng):
    for word, code in RAGWORD.items():
        f, fc = RAG[code]
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="equal", formula=[f'"{word}"'],
            fill=PatternFill("solid", fgColor=f),
            font=Font(name="Segoe UI", bold=True, color=fc)))


def navy_header(ws, row, names, widths):
    for j, n in enumerate(names, 1):
        c = ws.cell(row, j, n)
        c.font = HDR; c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = ctr; c.border = bd
        ws.column_dimensions[get_column_letter(j)].width = widths[j - 1]


def main():
    line = read_listing()
    last_row = {na: 2 + len(line[na]) for na in LINE_NAS}
    rwb = load_workbook(REGISTER, data_only=True)["Register"]
    reg_rows = []
    r = 5
    while rwb.cell(r, 1).value not in (None, ""):
        reg_rows.append([rwb.cell(r, c).value for c in range(1, 16)])
        r += 1
    twb = load_workbook(TRACKER, data_only=True)["Tracker"]
    trk = []
    for r in range(8, 95):
        if twb.cell(r, 2).value is None:
            continue
        trk.append([twb.cell(r, c).value for c in range(1, 13)])

    wb = Workbook()

    # ---- Tracker ----
    tk = wb.active; tk.title = "Tracker"
    tk["A1"] = "LCC Parks 4090000 — NA Review programme tracker"
    tk["A1"].font = Font(name="Segoe UI", bold=True, size=13, color=NAVY)
    dash = [("Accounts in scope", "=COUNTA(B8:B94)", "0"),
            ("Total in scope ex-GST", "=SUM(D8:D94)", "#,##0"),
            ("Closed & reconciled", '=SUMIF(I8:I94,"Closed",D8:D94)', "#,##0"),
            ("Partial: $ untested", '=SUMIF(I8:I94,"Partial",K8:K94)', "#,##0"),
            ("Queued accounts", '=COUNTIF(I8:I94,"Queued")', "0"),
            ("Top-4 concentration",
             "=(LARGE(D8:D94,1)+LARGE(D8:D94,2)+LARGE(D8:D94,3)+LARGE(D8:D94,4))/SUM(D8:D94)", "0.0%")]
    col = 1
    for lab, f, fmt in dash:
        tk.cell(5, col, lab).font = BOLD
        vc = tk.cell(6, col, f); vc.font = Font(name="Segoe UI", size=11, color=NAVY); vc.number_format = fmt
        col += 2
    navy_header(tk, 7, ["Priority", "NA", "Account name", "Accum Actual", "Annual Budget",
                        "Variance $", "Var %", "Materiality", "Status", "Reviewed $", "Untested $", "Notes"],
                [9, 8, 30, 14, 14, 13, 8, 12, 10, 13, 13, 40])
    reg_row_of = {str(row[0]): 5 + i for i, row in enumerate(reg_rows)}
    r = 8
    for row in trk:
        na = str(row[1])
        for j, v in enumerate(row, 1):
            tk.cell(r, j, v).font = BASE
        tk.cell(r, 6, f"=E{r}-D{r}"); tk.cell(r, 7, f'=IF(E{r}=0,"",(E{r}-D{r})/E{r})')
        if na in LINE_NAS:
            tk.cell(r, 10, f"=Register!P{reg_row_of[na]}")
            tk.cell(r, 11, f"=D{r}-J{r}")
        for c, fmt in ((4, "#,##0.00"), (5, "#,##0.00"), (6, "#,##0.00"), (7, "0.0%"),
                       (10, "#,##0.00"), (11, "#,##0.00")):
            tk.cell(r, c).number_format = fmt
        for c in range(1, 13):
            tk.cell(r, c).border = bd
        r += 1
    tk.freeze_panes = "A8"; tk.auto_filter.ref = f"A7:L{r-1}"

    # ---- Register ----
    rg = wb.create_sheet("Register")
    rg["A1"] = "LCC Parks 4090000 — Account review register (four-limb)"
    rg["A1"].font = Font(name="Segoe UI", bold=True, size=13, color=NAVY)
    navy_header(rg, 4, ["Account", "Account name", "Section / PK in scope", "$ ex-GST", "$ incl GST",
                        "L1 PK", "L2 NA", "L3 Evid", "L4 Tax", "Overall", "Finding", "Open actions",
                        "Evidence location", "Jrnl ref", "Reviewed date", "Reviewed $", "In journal $"],
                [9, 28, 30, 12, 12, 8, 8, 8, 8, 9, 34, 34, 28, 18, 13, 13, 13])
    rr = 5
    for src in reg_rows:
        na = str(src[0]); sh = na if na in LINE_NAS else None; last = last_row.get(na)
        rg.cell(rr, 1, na); rg.cell(rr, 2, src[1]); rg.cell(rr, 3, src[2])
        if sh:
            def w2(c1, c2):
                # Per-line rule: GREEN only when every line in BOTH columns is G.
                # Any R -> RED; any A -> AMBER; both columns all-P -> PENDING;
                # any P alongside reviewed lines -> AMBER (never GREEN off a subset).
                r1, r2 = f"'{sh}'!{c1}3:{c1}{last}", f"'{sh}'!{c2}3:{c2}{last}"
                return (f'=IF(OR(COUNTIF({r1},"R"),COUNTIF({r2},"R")),"RED",'
                        f'IF(OR(COUNTIF({r1},"A"),COUNTIF({r2},"A")),"AMBER",'
                        f'IF(AND(COUNTIF({r1},"P")=COUNTA({r1}),COUNTIF({r2},"P")=COUNTA({r2})),"PENDING",'
                        f'IF(COUNTIF({r1},"P")+COUNTIF({r2},"P")>0,"AMBER","GREEN"))))')

            def w1(c1):
                # Same rule, single column: GREEN only when every line is G.
                r1 = f"'{sh}'!{c1}3:{c1}{last}"
                return (f'=IF(COUNTIF({r1},"R"),"RED",'
                        f'IF(COUNTIF({r1},"A"),"AMBER",'
                        f'IF(COUNTIF({r1},"P")=COUNTA({r1}),"PENDING",'
                        f'IF(COUNTIF({r1},"P")>0,"AMBER","GREEN"))))')
            rg.cell(rr, 4, f"=SUM('{sh}'!K3:K{last})")
            rg.cell(rr, 6, w2("N", "O")); rg.cell(rr, 7, w1("P")); rg.cell(rr, 8, w1("Q"))
            rg.cell(rr, 9, w1("R"))
            rov = f"'{sh}'!S3:S{last}"
            rg.cell(rr, 10, (f'=IF(COUNTIF({rov},"R"),"RED",'
                             f'IF(COUNTIF({rov},"A"),"AMBER",'
                             f'IF(COUNTIF({rov},"P")=COUNTA({rov}),"PENDING",'
                             f'IF(COUNTIF({rov},"P")>0,"AMBER",'
                             f'IF(COUNTIF({rov},"J")>0,"CLEARED","GREEN")))))'))
            rg.cell(rr, 16, f"=SUMIFS('{sh}'!K3:K{last},'{sh}'!S3:S{last},\"<>P\")")
            rg.cell(rr, 17, f"=SUMIFS('{sh}'!K3:K{last},'{sh}'!S3:S{last},\"J\")")
        else:
            rg.cell(rr, 4, src[3])
            for j in range(6, 11):
                rg.cell(rr, j, src[j - 1])
            rg.cell(rr, 16, src[3] if src[9] != "PENDING" else 0)
            rg.cell(rr, 17, 0)
        if na == "73128":
            rg.cell(rr, 5, src[4])   # GST-free milk: invoice value, not D*1.1
        else:
            rg.cell(rr, 5, f"=D{rr}*1.1")
        for c, fmt in ((4, "#,##0.00"), (5, "#,##0.00"), (16, "#,##0.00"), (17, "#,##0.00")):
            rg.cell(rr, c).number_format = fmt
        rg.cell(rr, 11, src[10]); rg.cell(rr, 12, src[11]); rg.cell(rr, 13, src[12])
        rg.cell(rr, 14, src[13]); rg.cell(rr, 15, src[14])
        for c in range(1, 18):
            cell = rg.cell(rr, c); cell.font = BASE; cell.border = bd
            cell.alignment = ctr if c in (1, 4, 5, 6, 7, 8, 9, 10, 15, 16, 17) else lft
        rr += 1
    rg.cell(rr, 3, "TOTAL").font = BOLD
    rg.cell(rr, 4, f"=SUM(D5:D{rr-1})").number_format = "#,##0.00"
    rg.cell(rr, 5, f"=SUM(E5:E{rr-1})").number_format = "#,##0.00"
    rg.cell(rr, 16, f"=SUM(P5:P{rr-1})").number_format = "#,##0.00"
    rg.cell(rr, 17, f"=SUM(Q5:Q{rr-1})").number_format = "#,##0.00"
    for c in range(3, 18):
        rg.cell(rr, c).border = bd
    rag_cf(rg, f"F5:J{rr-1}")
    rg.freeze_panes = "A5"

    # ---- Summary + Legend ----
    def cnt(s, *c):
        return int(s.isin(c).sum())
    sm = wb.create_sheet("Summary")
    sm["A1"] = "Line-level review summary"; sm["A1"].font = Font(name="Segoe UI", bold=True, size=13, color=NAVY)
    navy_header(sm, 3, ["Account", "Lines", "$ ex-GST", "Reviewed %", "Overall R/A/P/J"], [40, 8, 14, 12, 18])
    sr = 4
    for na in LINE_NAS:
        df = line[na]; tot = pd.to_numeric(df["ExGST"], errors="coerce").sum()
        rev = int((df["Overall"] != "P").sum())
        for j, v in enumerate([na, len(df), round(tot, 2), f"{rev/len(df)*100:.0f}%",
                               f"{cnt(df['Overall'],'R')} / {cnt(df['Overall'],'A')} / {cnt(df['Overall'],'P')} / {cnt(df['Overall'],'J')}"], 1):
            c = sm.cell(sr, j, v); c.font = BASE; c.alignment = lft if j == 1 else ctr; c.border = bd
            if j == 3:
                c.number_format = "#,##0.00"
        sr += 1
    for label, note in [("73140 Pre-Employment Background Checks", "no line ledger; SE2 $3,703.03; see Verification Record"),
                        ("Remaining 80 programme accounts", "no TechOne transaction export pulled yet; line coverage pending")]:
        c = sm.cell(sr, 1, label); c.font = BASE; c.alignment = lft; c.border = bd
        c2 = sm.cell(sr, 5, note); c2.font = BASE; c2.alignment = lft; c2.border = bd
        sr += 1
    lg = wb.create_sheet("Legend")
    for i, (a, b) in enumerate([("Code", "Meaning"), ("G", "pass"), ("A", "confirm"),
                                ("R", "blocker / error"), ("P", "pending, not yet reviewed"),
                                ("J", "cleared - recode prepared and in the live journal; limb columns keep audit truth")], 1):
        ca = lg.cell(i, 1, a); cb = lg.cell(i, 2, b); ca.font = BOLD if i == 1 else BASE; cb.font = BASE
        if a in RAG:
            f, fc = RAG[a]; ca.fill = PatternFill("solid", fgColor=f)
            ca.font = Font(name="Segoe UI", bold=True, color=fc); ca.alignment = ctr
    lg.column_dimensions["A"].width = 8; lg.column_dimensions["B"].width = 40

    # ---- line tabs ----
    for na in LINE_NAS:
        df = line[na]; tot = pd.to_numeric(df["ExGST"], errors="coerce").sum()
        ws = wb.create_sheet(na)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(LCOLS))
        t = ws.cell(1, 1, f"NA {na} — line-level review ({len(df)} lines, ${tot:,.2f} ex-GST)")
        t.font = Font(name="Segoe UI", bold=True, size=12, color="FFFFFF")
        t.fill = PatternFill("solid", fgColor=NAVY)
        navy_header(ws, 2, LCOLS, LWID)
        r = 3
        for _, row in df.iterrows():
            for j, n in enumerate(LCOLS, 1):
                v = row[n]
                if n in RAGCOLS:
                    c = ws.cell(r, j, v); f, fc = RAG.get(v, ("FFFFFF", "000000"))
                    c.fill = PatternFill("solid", fgColor=f)
                    c.font = Font(name="Segoe UI", bold=True, color=fc, size=10); c.alignment = ctr
                else:
                    c = ws.cell(r, j, v); c.font = BASE
                    c.alignment = ctr if n in ("Period", "Date", "Attach", "ExGST", "InclX11", "Svc") else lft
                    if n in ("ExGST", "InclX11") and v not in (None, ""):
                        c.number_format = "#,##0.00"
                    if r % 2 == 0:
                        c.fill = band
                c.border = bd
            r += 1
        ws.freeze_panes = "G3"; ws.auto_filter.ref = f"A2:{get_column_letter(len(LCOLS))}{r-1}"

    wb.save("00_NAReview_Master.xlsx")
    print("built 00_NAReview_Master.xlsx; sheets:", wb.sheetnames)
    print("RECALC REQUIRED before retiring the three source files (formulas are uncached).")


if __name__ == "__main__":
    main()
